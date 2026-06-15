"""回款数据比对报告(第一期诊断,一次性离线脚本)。
逐项目比对"人工云文档(交付中心 xlsx) vs PMIS 系统",产出数据源归属清单+逐项目偏差+汇总。
只读,不接入管线、不改看板。依据 docs/superpowers/specs/2026-06-15-payment-compare-report-design.md。
"""
import csv
import os
import re
from datetime import datetime

# —— 阈值/常量(可调) ——
THRESH_NODE_RATIO = 0.01      # 节点级计划比例逐阶段差
THRESH_RATIO_PP = 0.10        # 项目级回款百分比差(10pp)
THRESH_AMOUNT = 50000.0       # 回款金额差(元)
THRESH_AMOUNT_REL = 0.20      # 回款金额相对差
THRESH_DAYS = 30              # 里程碑日期差(天)

CLOUD_XLSX = "input/交付中心-全量项目清单-交付实施三部.xlsx"
SHEET_MASTER = "数据源表_全量项目清单 (2)"
SHEET_NODES = "项目回款节点（里程碑）清单"
PMIS_DIR = "input/pmis"
INPUT_DIR = "input"
OUT_DIR = "report"
# PMIS 里程碑"关联回款阶段"列(阶段名→列名)
PMIS_STAGE_COLS = {"到货": "到货关联回款阶段", "初验": "初验关联回款阶段",
                   "终验": "终验关联回款阶段", "驻场": "驻场关联回款阶段"}


def parse_pay_stage_ratio(cell):
    """'到货款1，70.00%' → 0.70;空/无% → None。"""
    if not cell:
        return None
    m = re.search(r"([0-9]+(?:\.[0-9]+)?)\s*%", str(cell))
    return round(float(m.group(1)) / 100, 4) if m else None


def parse_ratio(v):
    """手填回款比例 → 0-1 小数。兼容 0.6 / '60%' / '60' / '' / None / '空值'。>1.5 视作百分数。"""
    if v is None or v == "" or v == "空值":
        return None
    if isinstance(v, (int, float)):
        return round(float(v), 4)
    s = str(v).strip().replace("%", "")
    if not s or s == "空值":
        return None
    try:
        n = float(s)
    except ValueError:
        return None
    return round(n / 100, 4) if n > 1.5 else round(n, 4)


def diff_flag(a, b, thresh):
    """|a-b| > thresh → True;任一 None → False(无法比,不算异常)。"""
    if a is None or b is None:
        return False
    return abs(a - b) > thresh


def node_actual_amount(project_amount, plan_ratio, actual_ratio):
    """节点实际回款金额 = 项目金额 × 关联回款比例 × 实际回款比例;任一 None → 0。"""
    return round((project_amount or 0) * (plan_ratio or 0) * (actual_ratio or 0), 2)


def days_between(d1, d2):
    """两个 'YYYY-MM-DD…' 字符串相差天数(绝对值);任一空/不可解析 → None。"""
    try:
        a = datetime.strptime(str(d1)[:10], "%Y-%m-%d")
        b = datetime.strptime(str(d2)[:10], "%Y-%m-%d")
        return abs((a - b).days)
    except (TypeError, ValueError):
        return None


def classify_level(strong_anomaly_count):
    return "红" if strong_anomaly_count >= 2 else ("黄" if strong_anomaly_count == 1 else "绿")


# —— 装载器 / 解析 / 比对 / 报告 / main ——
import warnings  # noqa: E402

import openpyxl  # noqa: E402

import milestones as milestones_mod  # noqa: E402
import pmis as pmis_mod  # noqa: E402
import profit as profit_mod  # noqa: E402
import projects as projects_mod  # noqa: E402

warnings.filterwarnings("ignore")


def _num(v):
    try:
        return round(float(v), 4)
    except (TypeError, ValueError):
        return None


def _col_idx(header, *subs):
    """按列名子串定位下标(首个命中);找不到 → None。"""
    for i, h in enumerate(header):
        hs = str(h or "").strip()
        if all(s in hs for s in subs):
            return i
    return None


def load_cloud_master(path):
    """主表 → {pid(当前在建项目编号): {项目名称,L4,合同总额,手填已回款比例,项目状态,合同编号}}。"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[SHEET_MASTER]
    it = ws.iter_rows(values_only=True)
    hdr = [str(h or "").strip() for h in next(it)]
    ci = {
        "pid": _col_idx(hdr, "当前在建项目编号"), "name": _col_idx(hdr, "项目名称"),
        "l4": _col_idx(hdr, "新L4组织"), "contract": _col_idx(hdr, "合同总额"),
        "ratio": _col_idx(hdr, "已回款比例"), "status": _col_idx(hdr, "项目状态"),
        "cno": _col_idx(hdr, "合同编号"), "pm": _col_idx(hdr, "项目经理"),
    }
    out = {}
    for r in it:
        pid = str(r[ci["pid"]]).strip() if ci["pid"] is not None and r[ci["pid"]] else ""
        if not pid:
            continue
        out[pid] = {
            "项目名称": r[ci["name"]] if ci["name"] is not None else "",
            "L4": r[ci["l4"]] if ci["l4"] is not None else "",
            "项目经理": r[ci["pm"]] if ci["pm"] is not None else "",
            "合同编号": r[ci["cno"]] if ci["cno"] is not None else "",
            "合同总额": r[ci["contract"]] if ci["contract"] is not None else None,
            "手填已回款比例": parse_ratio(r[ci["ratio"]]) if ci["ratio"] is not None else None,
            "项目状态": r[ci["status"]] if ci["status"] is not None else "",
        }
    wb.close()
    return out


def load_cloud_nodes(path):
    """节点清单 → {pid: {_amount, nodes: {阶段名: {plan_ratio, actual_ratio, plan_date, related}}}}。
    阶段名取自'里程碑节点'(到货/初验/终验/驻场…)。"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[SHEET_NODES]
    it = ws.iter_rows(values_only=True)
    hdr = [str(h or "").strip() for h in next(it)]
    ci = {k: _col_idx(hdr, s) for k, s in {
        "pid": "项目编号", "amount": "项目金额", "node": "里程碑节点",
        "plan_date": "该节点计划完成时间", "related": "是否关联回款",
        "plan_ratio": "关联回款比例", "actual_ratio": "实际回款比例"}.items()}
    out = {}
    for r in it:
        pid = str(r[ci["pid"]]).strip() if ci["pid"] is not None and r[ci["pid"]] else ""
        node = str(r[ci["node"]]).strip() if ci["node"] is not None and r[ci["node"]] else ""
        if not pid or not node:
            continue
        d = out.setdefault(pid, {"_amount": r[ci["amount"]] if ci["amount"] is not None else None, "nodes": {}})
        d["nodes"][node] = {
            "plan_ratio": _num(r[ci["plan_ratio"]]) if ci["plan_ratio"] is not None else None,
            "actual_ratio": _num(r[ci["actual_ratio"]]) if ci["actual_ratio"] is not None else None,
            "plan_date": r[ci["plan_date"]] if ci["plan_date"] is not None else None,
            "related": str(r[ci["related"]]).strip() == "是" if ci["related"] is not None and r[ci["related"]] else False,
        }
    wb.close()
    return out


def load_pmis_stage_ratios(pmis_dir):
    """读两张里程碑表(read_only=False,表头第2行),解析4个关联回款阶段列 →
    {pid: {阶段名: 计划回款比例}}。同 pid 在建优先(先读在建,已结项不覆盖)。"""
    import config
    out = {}
    for fn in [config.MILESTONE_FILE_ACTIVE, config.MILESTONE_FILE_CLOSED]:
        fp = os.path.join(pmis_dir, fn)
        if not os.path.exists(fp):
            continue
        wb = openpyxl.load_workbook(fp, read_only=False, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if len(rows) < 2:
            continue
        hdr = [str(h or "").strip() for h in rows[1]]          # 表头第2行
        pid_i = _col_idx(hdr, "项目编号")
        stage_i = {st: _col_idx(hdr, col) for st, col in PMIS_STAGE_COLS.items()}
        for r in rows[2:]:
            if not any(r):
                continue
            pid = str(r[pid_i]).strip() if pid_i is not None and pid_i < len(r) and r[pid_i] else ""
            if not pid or pid in out:                          # 在建优先:已存在则跳过
                continue
            d = {}
            for st, ci in stage_i.items():
                if ci is not None and ci < len(r):
                    rt = parse_pay_stage_ratio(r[ci])
                    if rt is not None:
                        d[st] = rt
            if d:
                out[pid] = d
    return out


def build_rows(base_dir):
    master = load_cloud_master(os.path.join(base_dir, CLOUD_XLSX))
    nodes = load_cloud_nodes(os.path.join(base_dir, CLOUD_XLSX))
    # 售前→原项目映射(A.xlsx):售前项目自身无 PMIS 合同/流水/里程碑,回退到原项目(已关闭)检索
    origin = {mm["current"]: mm["closed"]
              for mm in projects_mod.read_mapping(os.path.join(base_dir, INPUT_DIR, "A.xlsx"))}
    closed_ids = set(origin.values())
    keep = set(master.keys())
    keep_ext = keep | closed_ids
    pay, _ = profit_mod.load_payment_records(os.path.join(base_dir, INPUT_DIR), keep_ext)
    pmis_map, _ = pmis_mod.load_project_pmis(os.path.join(base_dir, PMIS_DIR), keep, extra_closed_ids=closed_ids)
    pmis_stage = load_pmis_stage_ratios(os.path.join(base_dir, PMIS_DIR))   # 全量,自带原项目
    ms, _, _ = milestones_mod.load_milestones(os.path.join(base_dir, PMIS_DIR), keep_ext)
    rows = []
    for pid, m in master.items():
        oid = origin.get(pid, "")
        used_origin = False
        # 合同总额:本项目 PMIS 优先;售前等自身无则回退原项目(oid);再缺回退云文档主表(spec §4)
        pmis_contract = _num(((pmis_map.get(pid) or {}).get("customer") or {}).get("合同总额"))
        if not pmis_contract and oid:
            oc = _num(((pmis_map.get(oid) or {}).get("customer") or {}).get("合同总额"))
            if oc:
                pmis_contract, used_origin = oc, True
        contract = pmis_contract if pmis_contract else _num(m["合同总额"])
        # 流水:本项目优先,自身无则回退原项目
        rec = pay.get(pid) or {}
        if not rec.get("total") and oid and (pay.get(oid) or {}).get("total"):
            rec, used_origin = pay.get(oid), True
        pmis_total = rec.get("total")
        pmis_count = rec.get("count", 0)
        pmis_ratio = round(pmis_total / contract, 4) if (pmis_total is not None and contract) else None
        # 项目回款百分比
        ratio_diff = (None if (pmis_ratio is None or m["手填已回款比例"] is None)
                      else round(pmis_ratio - m["手填已回款比例"], 4))
        ratio_anom = diff_flag(pmis_ratio, m["手填已回款比例"], THRESH_RATIO_PP)
        # 节点级计划比例逐阶段分歧(里程碑侧本项目优先,售前回退原项目)
        cnodes = (nodes.get(pid) or {}).get("nodes", {})
        camount = (nodes.get(pid) or {}).get("_amount")
        eff = pid
        if not pmis_stage.get(pid) and oid and pmis_stage.get(oid):
            eff, used_origin = oid, True
        pstage = pmis_stage.get(eff, {})
        ratio_mismatch_stages = []
        for st, prt in pstage.items():
            crt = next((v["plan_ratio"] for k, v in cnodes.items() if st in k), None)
            if crt is not None and diff_flag(prt, crt, THRESH_NODE_RATIO):
                ratio_mismatch_stages.append(f"{st}(PMIS{prt}/云{crt})")
        node_ratio_anom = bool(ratio_mismatch_stages)
        # 回款金额:云节点 Σ实际 vs PMIS 流水累计
        cloud_actual_amt = round(sum(
            node_actual_amount(camount, v["plan_ratio"], v["actual_ratio"])
            for v in cnodes.values() if v["related"]), 2)
        amt_diff = (None if pmis_total is None else round(pmis_total - cloud_actual_amt, 2))
        amt_anom = diff_flag(pmis_total, cloud_actual_amt, THRESH_AMOUNT)
        # 笔数/个数
        cloud_paid_nodes = sum(1 for v in cnodes.values() if v["related"] and (v["actual_ratio"] or 0) >= 1)
        cloud_rel_nodes = sum(1 for v in cnodes.values() if v["related"])
        pmis_rel_ms = len(pstage)
        # 里程碑日期分歧(按阶段对齐 PMIS 里程碑 planDate)
        date_anom_stages = []
        for st in pstage:
            c_node = next((v for k, v in cnodes.items() if st in k), None)
            p_ms = next((it_ for key, it_ in ((i.get("name"), i) for i in (ms.get(eff) or []))
                         if key and st in str(key)), None)
            if c_node and p_ms:
                dd = days_between(c_node["plan_date"], p_ms.get("planDate"))
                if dd is not None and dd > THRESH_DAYS:
                    date_anom_stages.append(f"{st}({dd}天)")
        strong = sum([node_ratio_anom, ratio_anom, amt_anom, bool(date_anom_stages)])
        flags = []
        if m["手填已回款比例"] is None:
            flags.append("手填比例未填")
        if pmis_total is None:
            flags.append("PMIS无流水")
        if not pmis_contract:
            flags.append("PMIS无合同总额")
        if not pstage:
            flags.append("PMIS无关联回款比例")
        if cloud_rel_nodes != pmis_rel_ms:
            flags.append("节点个数不一致")
        if cloud_paid_nodes != pmis_count:
            flags.append("笔数不一致")
        if used_origin:
            flags.append("售前取原项目")
        rows.append({
            "pid": pid, "项目名称": m["项目名称"], "L4": m["L4"], "项目经理": m["项目经理"],
            "原项目编号": oid if used_origin else "",
            "合同编号": m["合同编号"], "合同总额": contract, "项目状态": m["项目状态"],
            "手填已回款比例": m["手填已回款比例"], "PMIS流水比例": pmis_ratio, "比例差": ratio_diff,
            "比例异常": ratio_anom, "节点比例分歧阶段": "；".join(ratio_mismatch_stages),
            "云已回款金额": cloud_actual_amt, "PMIS流水累计": pmis_total, "金额差": amt_diff, "金额异常": amt_anom,
            "云已回款节点数": cloud_paid_nodes, "PMIS笔数": pmis_count,
            "云回款节点数": cloud_rel_nodes, "PMIS关联回款里程碑数": pmis_rel_ms,
            "里程碑日期异常阶段": "；".join(date_anom_stages),
            "强异常维度数": strong, "等级": classify_level(strong), "标记位": "；".join(flags),
        })
    return rows, {"master": len(master), "pay": len(pay), "pmis_stage": len(pmis_stage)}


CSV_COLS = ["pid", "项目名称", "L4", "项目经理", "原项目编号", "合同编号", "合同总额", "项目状态",
            "手填已回款比例", "PMIS流水比例", "比例差", "比例异常", "节点比例分歧阶段",
            "云已回款金额", "PMIS流水累计", "金额差", "金额异常",
            "云已回款节点数", "PMIS笔数", "云回款节点数", "PMIS关联回款里程碑数",
            "里程碑日期异常阶段", "强异常维度数", "等级", "标记位"]


def write_reports(rows, stat, out_dir):
    os.makedirs(out_dir, exist_ok=True)
    # 逐项目 CSV
    with open(os.path.join(out_dir, "回款比对_逐项目.csv"), "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=CSV_COLS, extrasaction="ignore")
        w.writeheader()
        for r in sorted(rows, key=lambda x: -x["强异常维度数"]):
            w.writerow(r)
    # 汇总 MD
    n = len(rows)
    red = sum(1 for r in rows if r["等级"] == "红")
    yellow = sum(1 for r in rows if r["等级"] == "黄")
    from collections import Counter
    fc = Counter(fl for r in rows for fl in (r["标记位"].split("；") if r["标记位"] else []))
    lines = ["# 回款比对汇总", "",
             f"- 参与比对项目: {n}（红 {red} / 黄 {yellow} / 绿 {n-red-yellow}）",
             f"- join 命中: 主表 {stat['master']} / 流水 {stat['pay']} / PMIS含比例 {stat['pmis_stage']}", "",
             "## 标记位分布"]
    for k, v in fc.most_common():
        lines.append(f"- {k}: {v}（{round(v / n * 100, 1) if n else 0.0}%）")
    lines += ["", "## 红等级 Top 异常"]
    for r in [r for r in rows if r["等级"] == "红"][:50]:
        lines.append(f"- {r['pid']} {r['项目名称']}: 比例差 {r['比例差']} / 金额差 {r['金额差']} / "
                     f"节点分歧 {r['节点比例分歧阶段']} / 日期异常 {r['里程碑日期异常阶段']}")
    lines += ["", "## 阈值", f"- 节点比例 {THRESH_NODE_RATIO} / 百分比 {THRESH_RATIO_PP} / "
              f"金额 {THRESH_AMOUNT} / 日期 {THRESH_DAYS}天"]
    with open(os.path.join(out_dir, "回款比对_汇总.md"), "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    # 数据源归属清单(静态结论,见 spec §3)
    with open(os.path.join(out_dir, "数据源归属清单.md"), "w", encoding="utf-8") as f:
        f.write(SRC_INVENTORY_MD)


SRC_INVENTORY_MD = '''# 数据源归属清单(回款看板字段 → PMIS可/人工only)

## PMIS 可（以 PMIS 为准）
项目编号/金额/合同总额、tier、orgL4/项目经理、里程碑节点名+计划/实际日期、
**节点级计划回款比例**(解析里程碑"关联回款阶段")、是否已达成里程碑(actualDate非空近似)、
项目实际到账(流水累计)、回款笔数、项目回款%(流水÷合同)、签约形式分类。

## 人工 only（无法 PMIS 化）
- G1 节点级实际回款比例(严格手填;可用里程碑达成+项目流水重定义)
- G2 加资源可提前 canAdvance
- G3 业务分类标签(BH/退换货/0元单/框架/维保;"已100%回款""已关闭"可PMIS派生)
- G4 纳管(工具自有域控制开关)
- G5 跟进记录(工具自有+云文档回写)

> G3/G4 第二期由项目标签体系承载;G1 第二期改 PMIS 口径;G2 多半可弃。
'''


def main():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    rows, stat = build_rows(base_dir)
    write_reports(rows, stat, os.path.join(base_dir, OUT_DIR))
    red = sum(1 for r in rows if r["等级"] == "红")
    print(f"[OK] 比对完成: {len(rows)} 项目, 红 {red}; 报告见 {OUT_DIR}/")


if __name__ == "__main__":
    main()
