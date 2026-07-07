# -*- coding: utf-8 -*-
"""projects.py 纯函数单元测试。不依赖 input/ 真文件——用 tmp_path 生成的小 xlsx 或内存 dict。"""
import openpyxl
import pytest

import config
import projects as P


def _make_xlsx(dir_path, name, sheets):
    """造多 sheet xlsx。sheets = [(sheet名, rows[list[tuple]])]，首个 sheet 复用默认 active。"""
    wb = openpyxl.Workbook()
    for i, (title, rows) in enumerate(sheets):
        ws = wb.active if i == 0 else wb.create_sheet()
        ws.title = title
        for r in rows:
            ws.append(list(r))
    path = str(dir_path / name)
    wb.save(path)
    return path


class TestReadOrgNames:
    def test_picks_sheet_with_gonghao_header(self, tmp_path):
        path = _make_xlsx(tmp_path, "组织架构.xlsx", [
            ("Sheet2", [("行标签", "求和项:成本"), ("银行服务组", 100)]),  # 透视杂表在前
            ("Sheet1", [
                ("工号", "姓名", "员工类别", "新L2组织", "新L3组织", "新L3-1组织", "新L4组织",
                 "直接上级工号", "直接上级姓名", "是否项目经理", "成本"),
                ("A012804", "佘海龙", "正式员工", "交付中心", "交付实施三部", "服务二部",
                 "黑龙江服务组", "A001373", "于岩", None, 1500),
                ("A002338", "杨亮", "正式员工", "交付中心", "交付实施三部", "服务二部",
                 "黑龙江服务组", "A012804", "佘海龙", None, 1000),
            ]),
        ])
        names, l4s, rows = P.read_org_names(path)
        assert names == {"佘海龙", "杨亮"}
        assert l4s == {"黑龙江服务组"}
        assert rows == 2

    def test_missing_file_degrades(self, tmp_path):
        names, l4s, rows = P.read_org_names(str(tmp_path / "不存在.xlsx"))
        assert names == set() and l4s == set() and rows == 0


class TestReadMapping:
    def test_headerless_three_cols(self, tmp_path):
        path = _make_xlsx(tmp_path, "A.xlsx", [
            ("Sheet1", [
                ("WSGF-SF-202301100425", "于江", "WSGF-SS-202212229197"),
                ("WSGF-SF-202304190139", "于江", "WSGF-SS-202303289058"),
                (None, None, None),  # 空行跳过
                ("WSGF-SF-X", "某人", None),  # 缺已关闭号跳过
            ]),
        ])
        m = P.read_mapping(path)
        assert m == [
            {"current": "WSGF-SF-202301100425", "owner": "于江", "closed": "WSGF-SS-202212229197"},
            {"current": "WSGF-SF-202304190139", "owner": "于江", "closed": "WSGF-SS-202303289058"},
        ]

    def test_missing_file_degrades(self, tmp_path):
        assert P.read_mapping(str(tmp_path / "无.xlsx")) == []


class TestReadDelivery:
    def test_picks_sheet_with_pid_header_and_skips_pivot(self, tmp_path):
        path = _make_xlsx(tmp_path, "delivery_analysis.xlsx", [
            ("Sheet1", [(None, None), ("行标签", "求和项:x")]),  # 透视杂表
            ("delivery_analysis (1)", [
                ("项目编号", "项目名称", "项目经理", "L4组织", "差旅费_预算金额", "差旅费_消耗率"),
                ("WSGF-SS-1", "某项目", "佘海龙", "黑龙江服务组", 1000, "50%"),
            ]),
        ])
        rows = P.read_delivery(path)
        assert len(rows) == 1
        assert rows[0]["项目编号"] == "WSGF-SS-1"
        assert rows[0]["差旅费_预算金额"] == 1000


class TestReadDeliveryCsv:
    def test_csv_first_and_legacy_fallback(self, tmp_path):
        import projects as PJ
        csv_path = tmp_path / "delivery_analysis.csv"
        csv_path.write_text("项目编号,项目名称,交付外包服务成本_预算金额\nSS-1,甲,100.0\n", encoding="utf-8-sig")
        rows = PJ.read_delivery(str(csv_path))
        assert rows[0]["项目编号"] == "SS-1"
        # csv 缺失 → 回退 xlsx(不存在则空)
        assert PJ.read_delivery(str(tmp_path / "none" / "delivery_analysis.csv")) == []


class TestDeliveryCostsFor:
    def test_seven_categories_parsed(self):
        row = {"差旅费_预算金额": "1,000", "差旅费_实际发生": 600,
               "差旅费_剩余预算": 400, "差旅费_消耗率": "60%"}
        out = P.delivery_costs_for(row)
        assert len(out) == len(config.DELIVERY_COST_CATEGORIES)
        trip = next(i for i in out if i["类别"] == "差旅费")
        assert trip == {"类别": "差旅费", "预算金额": 1000.0, "实际发生": 600.0,
                        "剩余预算": 400.0, "消耗率": pytest.approx(0.6)}
        other = next(i for i in out if i["类别"] == "其他费用")
        assert other["预算金额"] is None  # 缺列降 None


class TestComputeHealth:
    def _pm(self, **over):
        pm = {"progress": {"里程碑进度状态": "正常"},
              "risk": {"最高等级": "低", "未关闭风险数": 0},
              "cost": {"项目超支": False, "消耗比": 0.5}}
        pm.update(over)
        return pm

    def test_all_ok(self):
        h = P.compute_health(self._pm(), delayed_count=0)
        assert h["overall"] == "健康"
        assert not any([h["progressAbnormal"], h["riskAbnormal"],
                        h["costAbnormal"], h["paymentAbnormal"]])

    def test_one_abnormal_is_warn(self):
        h = P.compute_health(self._pm(progress={"里程碑进度状态": "严重延期"}), 0)
        assert h["progressAbnormal"] is True and h["overall"] == "关注"

    def test_progress_abnormal_overdue_unpublished(self):
        h = P.compute_health(self._pm(progress={"里程碑进度状态": "超期未发布"}), 0)
        assert h["progressAbnormal"] is True

    def test_empty_pm_degrades_to_healthy(self):
        h = P.compute_health({}, 0)
        assert h["overall"] == "健康"
        assert not any([h["progressAbnormal"], h["riskAbnormal"],
                        h["costAbnormal"], h["paymentAbnormal"]])

    def test_two_abnormal_is_risk(self):
        h = P.compute_health(self._pm(risk={"最高等级": "高", "未关闭风险数": 2}), 1)
        assert h["riskAbnormal"] and h["paymentAbnormal"] and h["overall"] == "风险"

    def test_cost_abnormal_by_ratio_or_overrun(self):
        assert P.compute_health(self._pm(cost={"项目超支": True, "消耗比": 0.2}), 0)["costAbnormal"]
        assert P.compute_health(self._pm(cost={"项目超支": None, "消耗比": 1.2}), 0)["costAbnormal"]
        assert not P.compute_health(self._pm(cost={"项目超支": None, "消耗比": None}), 0)["costAbnormal"]


def _pm_active(name, manager, l4="黑龙江服务组", project_type="实施项目", l3_1="三部一组", **over):
    pm = {"matched": True, "source": "在建",
          "team": {"项目名称": name, "项目经理": manager, "L4部门": l4, "L3_1部门": l3_1},
          "progress": {"里程碑进度状态": "正常"},
          "risk": {"最高等级": None, "未关闭风险数": 0},
          "cost": {"项目超支": None, "消耗比": None},
          "status": {"项目类型": project_type},
          "customer": {"合同编号": "HT-" + (name or "x")}}
    pm.update(over)
    return pm


class TestBuildProjects:
    def test_filters_active_and_dept(self):
        ppm = {
            "SF-1": _pm_active("售前服务A", "佘海龙"),
            "SS-9": _pm_active("外部项目", "外部人"),       # 经理不在清单 → 排除
            "SS-8": {**_pm_active("已关闭项目", "佘海龙"), "source": "已关闭"},  # 非在建 → 排除
        }
        out = P.build_projects(ppm, {"佘海龙"}, {"黑龙江服务组"}, [], [])
        assert [p["projectId"] for p in out] == ["SF-1"]

    def test_org_missing_degrades_to_all_active(self):
        ppm = {"SS-1": _pm_active("某项目", "任意人")}
        out = P.build_projects(ppm, set(), set(), [], [])
        assert len(out) == 1  # 空人员清单=不过滤(spec 3.4 降级)

    def test_presale_mapping_and_payment(self):
        ppm = {"SF-1": _pm_active("售前服务A", "佘海龙", project_type="售前服务类")}
        mapping = [{"current": "SF-1", "owner": "于江", "closed": "SS-99"}]
        delivery = [{"项目编号": "SF-1", "项目名称": "售前服务A", "差旅费_预算金额": 100}]
        out = P.build_projects(ppm, {"佘海龙"}, {"黑龙江服务组"}, mapping, delivery)
        p = out[0]
        assert p["isPresale"] is True
        assert p["relatedClosedId"] == "SS-99"
        assert p["orgL3_1"] == "三部一组"
        assert p["合同编号"] == "HT-售前服务A"
        assert "orgL3" not in p
        assert next(i for i in p["deliveryCosts"] if i["类别"] == "差旅费")["预算金额"] == 100.0

    def test_name_falls_back_to_nodes(self):
        # nodes 名称回填已移除，name 来自 PMIS team；空名称项目 projectName 为空字符串
        ppm = {"SS-1": _pm_active(None, "佘海龙")}
        out = P.build_projects(ppm, {"佘海龙"}, set(), [], [])
        assert out[0]["projectName"] == ""

    def test_unmatched_pm_health_no_data(self):
        ppm = {"SS-1": {**_pm_active("某项目", "佘海龙"), "matched": False}}
        out = P.build_projects(ppm, {"佘海龙"}, set(), [], [])
        assert out[0]["health"]["overall"] == "无数据"
        assert out[0]["health"]["paymentAbnormal"] is False


class TestProjectsQuality:
    def test_quality_counts_and_alerts(self):
        ppm = {
            "SF-1": _pm_active("售前服务A", "佘海龙", project_type="售前服务类"),
            "SS-2": _pm_active("漏网项目", "王漏网", l4="黑龙江服务组"),  # L4 命中但经理不在清单 → 告警
        }
        projects = P.build_projects(ppm, {"佘海龙", "杨亮"}, {"黑龙江服务组"},
                                    [{"current": "SF-1", "owner": "x", "closed": "SS-99"}],
                                    [{"项目编号": "SF-1"}])
        q = P.compute_projects_quality(projects, ppm, {"佘海龙", "杨亮"}, {"黑龙江服务组"}, 2,
                                       [{"current": "SF-1", "owner": "x", "closed": "SS-99"}],
                                       [{"项目编号": "SF-1"}, {"项目编号": "SS-外部"}])
        assert q["deptProjectCount"] == 1
        assert q["staffNoProject"] == [{"name": "杨亮"}]
        assert q["managerNotInOrg"] == [{"projectId": "SS-2", "projectName": "漏网项目",
                                         "manager": "王漏网"}]
        assert q["presaleTotal"] == 1 and q["presaleMapped"] == 1 and q["presaleUnmapped"] == []
        assert q["mappingFile"] == {"provided": True, "rows": 1, "matched": 1, "matchRate": 1.0}
        assert q["deliveryFile"]["matched"] == 1 and q["deliveryFile"]["rows"] == 2


class TestReadOrgNamesEnhanced:
    def test_whitespace_name_skipped_and_other_dept_filtered(self, tmp_path):
        path = _make_xlsx(tmp_path, "组织架构.xlsx", [
            ("Sheet1", [
                ("工号", "姓名", "新L3组织", "新L4组织"),
                ("A1", "佘海龙", "交付实施三部", "黑龙江服务组"),
                ("A2", " ", "交付实施三部", "黑龙江服务组"),      # 空白姓名跳过
                ("A3", "外部门人", "交付实施一部", "别的组"),      # 非三部行过滤
            ]),
        ])
        names, l4s, rows = P.read_org_names(path)
        assert names == {"佘海龙"}
        assert l4s == {"黑龙江服务组"}
        assert rows == 2  # 三部行数(含空白姓名行)

    def test_no_l3_column_no_filter(self, tmp_path):
        path = _make_xlsx(tmp_path, "组织架构.xlsx", [
            ("Sheet1", [("工号", "姓名", "新L4组织"), ("A1", "某人", "某组")]),
        ])
        names, _, rows = P.read_org_names(path)
        assert names == {"某人"} and rows == 1


class TestReadDegradation:
    def test_header_not_found_in_any_sheet(self, tmp_path):
        path = _make_xlsx(tmp_path, "x.xlsx", [("Sheet1", [("无关", "列"), ("a", 1)])])
        assert P.read_delivery(path) == []

    def test_corrupt_file_degrades(self, tmp_path):
        bad = tmp_path / "坏.xlsx"
        bad.write_bytes(b"not an xlsx at all")
        assert P.read_delivery(str(bad)) == []
        names, l4s, rows = P.read_org_names(str(bad))
        assert names == set() and rows == 0


class TestDeliveryOverspendCats:
    def test_over_categories_listed(self):
        import projects as PJ
        costs = [
            {"类别": "交付外包服务成本", "预算金额": 100.0, "实际发生": 150.0},
            {"类别": "差旅费", "预算金额": 200.0, "实际发生": 100.0},
            {"类别": "项目直接成本", "预算金额": 0.0, "实际发生": 50.0},   # 预算0实际>0 也算超
            {"类别": "其他", "预算金额": None, "实际发生": 10.0},          # 预算缺失不算
        ]
        assert PJ.delivery_overspend_cats(costs) == ["交付外包服务成本", "项目直接成本"]
        assert PJ.delivery_overspend_cats([]) == []


class TestPaymentRatioFromRecords:
    def test_normal_and_presale_fallback(self):
        import projects as PJ
        assert PJ.payment_ratio_from_records(500.0, 1000.0, None) == 0.5
        assert PJ.payment_ratio_from_records(1151500.0, None, 1151500.0) == 1.0   # 售前取原项目
        assert PJ.payment_ratio_from_records(None, 1000.0, None) == 0.0           # 无流水=0%
        assert PJ.payment_ratio_from_records(500.0, None, None) is None           # 分母缺失
        assert PJ.payment_ratio_from_records(500.0, 0, 0) is None


class TestBuildPaymentSummary:
    def _node(self, stage, expected, reached, status):
        return {"stage": stage, "expectedPayment": expected, "reached": reached, "status": status}

    def test_summary_from_nodes(self):
        import projects as PJ
        nodes = [self._node("到货款", 700000.0, True, "已回款"),
                 self._node("终验款", 300000.0, False, "延期")]
        rec = {"total": 700000.0, "count": 2, "lastDate": "2026-06-04"}
        s = PJ.build_payment_summary(1000000.0, nodes, rec)
        assert s["contract"] == 1000000.0 and s["actualTotal"] == 700000.0 and s["paymentCount"] == 2
        assert "paymentRatio" not in s
        assert s["expectedTotal"] == 1000000.0
        assert s["nodeCount"] == 2 and s["reachedCount"] == 1 and s["delayedCount"] == 1
        assert s["lastPaymentDate"] == "2026-06-04" and s["fromOrigin"] is False

    def test_robust_none(self):
        import projects as PJ
        s = PJ.build_payment_summary(None, [], None)
        assert "paymentRatio" not in s and s["actualTotal"] is None
        assert s["expectedTotal"] == 0 and s["nodeCount"] == 0
        assert s["reachedCount"] == 0 and s["delayedCount"] == 0


class TestOrgL31AndContract:
    def test_build_projects_sets_orgL3_1_and_contract(self):
        pmis = {"P1": {"source": "在建", "matched": True,
                       "team": {"项目经理": "张三", "项目名称": "甲", "L4部门": "北京服务组",
                                "L3_1部门": "三部一组"},
                       "status": {"项目类型": "实施项目"},
                       "customer": {"合同编号": "HT-1"}}}
        projs = P.build_projects(pmis, {"张三"}, {"北京服务组"}, [], [])
        assert projs[0]["orgL3_1"] == "三部一组"
        assert projs[0]["合同编号"] == "HT-1"
        assert projs[0]["isPresale"] is False

    def test_isPresale_by_project_type(self):
        pmis = {"P1": {"source": "在建", "matched": True,
                       "team": {"项目经理": "张三", "项目名称": "未命名", "L4部门": "北京服务组",
                                "L3_1部门": "组"},
                       "status": {"项目类型": "售前服务类"}, "customer": {}}}
        projs = P.build_projects(pmis, {"张三"}, set(), [], [])
        assert projs[0]["isPresale"] is True


class TestCountClosedDept:
    def test_counts_manager_in_org(self, tmp_path):
        import openpyxl, os as _os
        d = tmp_path / "pmis"; d.mkdir()
        wb = openpyxl.Workbook(); ws = wb.active
        ws.cell(row=1, column=1, value="标题")
        ws.cell(row=2, column=1, value="项目编号"); ws.cell(row=2, column=2, value="项目经理")
        ws.cell(row=3, column=1, value="C-1"); ws.cell(row=3, column=2, value="张三")
        ws.cell(row=4, column=1, value="C-2"); ws.cell(row=4, column=2, value="外部人")
        wb.save(str(d / config.PMIS_FILES_CLOSED["center"]))
        assert P.count_closed_dept(str(d), {"张三"}) == 1
        assert P.count_closed_dept(str(d), set()) == 0
        assert P.count_closed_dept(str(tmp_path / "none"), {"张三"}) == 0


class TestAggregatePaymentPmis:
    def test_node_level(self):
        nodes = [
            {"expectedPayment": 1000000, "receivedAmount": 600000, "unpaidAmount": 400000, "status": "部分回款"},
            {"expectedPayment": 1000000, "receivedAmount": 0, "unpaidAmount": 1000000, "status": "延期"},
        ]
        r = P.aggregate_payment_pmis(nodes)
        assert r["relatedNodeCount"] == 2
        assert r["expectedTotal"] == 2000000
        assert r["actualTotal"] == 600000
        assert r["remainingTotal"] == 1400000
        assert r["paymentRatio"] is None
        assert r["delayedCount"] == 1
    def test_empty(self):
        r = P.aggregate_payment_pmis([])
        assert r["relatedNodeCount"] == 0 and r["paymentRatio"] is None


def test_aggregate_payment_pmis_ratio_is_none():
    import projects
    nodes = [{'expectedPayment': 100, 'receivedAmount': 50, 'unpaidAmount': 50, 'status': '部分回款', 'reached': False}]
    assert projects.aggregate_payment_pmis(nodes)['paymentRatio'] is None


class TestReadTop1000:
    def test_parses_name_level_quadrant_and_strips(self, tmp_path):
        path = _make_xlsx(tmp_path, "TOP1000.xlsx", [
            ("Sheet1", [
                ("客户编码", "客户名称", "客户级别", "象限"),
                ("C001", "辽宁省公安厅", "TOP1000大客户", "M1 战略核心区"),
                ("C002", " 北京能源集团 ", "TOP1000大客户", " M1 战略核心区 "),
            ]),
        ])
        m = P.read_top1000(path)
        assert m["辽宁省公安厅"] == {"level": "TOP1000大客户", "quad": "M1 战略核心区"}
        assert m["北京能源集团"] == {"level": "TOP1000大客户", "quad": "M1 战略核心区"}

    def test_skips_empty_name_rows(self, tmp_path):
        path = _make_xlsx(tmp_path, "TOP1000.xlsx", [
            ("Sheet1", [
                ("客户编码", "客户名称", "客户级别", "象限"),
                ("C001", None, "TOP1000大客户", "M1 战略核心区"),
                ("C002", "有名客户", "TOP1000大客户", "M2 现金牛/打猎区"),
            ]),
        ])
        m = P.read_top1000(path)
        assert list(m.keys()) == ["有名客户"]

    def test_missing_file_degrades_to_empty(self, tmp_path):
        assert P.read_top1000(str(tmp_path / "无.xlsx")) == {}


class TestBuildProjectsTop1000:
    def _ppm(self, final_customer):
        pm = _pm_active("项目甲", "佘海龙")
        pm["customer"]["最终客户"] = final_customer
        return {"SS-1": pm}

    def test_matched_top1000_level_yes_with_quadrant(self):
        m = {"辽宁省公安厅": {"level": "TOP1000大客户", "quad": "M1 战略核心区"}}
        out = P.build_projects(self._ppm("辽宁省公安厅"), {"佘海龙"}, set(), [], [], m)
        assert out[0]["top1000"] == "是"
        assert out[0]["quadrant"] == "M1 战略核心区"

    def test_matched_non_top1000_level_is_no_but_quadrant_kept(self):
        m = {"某客户": {"level": "TOP1001大客户", "quad": "M2 现金牛/打猎区"}}
        out = P.build_projects(self._ppm("某客户"), {"佘海龙"}, set(), [], [], m)
        assert out[0]["top1000"] == "否"
        assert out[0]["quadrant"] == "M2 现金牛/打猎区"

    def test_unmatched_is_no_empty_quadrant(self):
        m = {"辽宁省公安厅": {"level": "TOP1000大客户", "quad": "M1 战略核心区"}}
        out = P.build_projects(self._ppm("不在表里"), {"佘海龙"}, set(), [], [], m)
        assert out[0]["top1000"] == "否"
        assert out[0]["quadrant"] == ""

    def test_no_map_degrades_to_no(self):
        out = P.build_projects(self._ppm("辽宁省公安厅"), {"佘海龙"}, set(), [], [])
        assert out[0]["top1000"] == "否"
        assert out[0]["quadrant"] == ""

    def test_empty_final_customer_never_matches_even_if_map_has_empty_key(self):
        # 纵深防御:最终客户为空时,即便 map 意外含空键 "" 也不得命中
        m = {"": {"level": "TOP1000大客户", "quad": "M1 战略核心区"}}
        out = P.build_projects(self._ppm(""), {"佘海龙"}, set(), [], [], m)
        assert out[0]["top1000"] == "否"
        assert out[0]["quadrant"] == ""


class TestBuildProjectsPresaleTop1000:
    """售前服务类:TOP1000/象限 按原项目(relatedClosedId)最终客户判定,不用本项目最终客户。
    售前本项目最终客户实测全为空;无原项目/原项目无客户 → 否/空(不回退本项目)。"""
    TOPMAP = {"辽宁省公安厅": {"level": "TOP1000大客户", "quad": "M1 战略核心区"}}

    def _ppm(self, own_customer, orig_customer, orig_in_pm=True):
        sf = _pm_active("售前甲", "佘海龙", project_type="售前服务类")
        sf["customer"]["最终客户"] = own_customer
        ppm = {"SF-1": sf}
        if orig_in_pm:
            ss = {**_pm_active("原项目甲", "佘海龙"), "source": "已关闭"}
            ss["customer"]["最终客户"] = orig_customer
            ppm["SS-99"] = ss
        return ppm

    def test_presale_judged_by_original_customer(self):
        # 本项目客户空,原项目客户在清单 → 是 + 原项目象限
        ppm = self._ppm(own_customer="", orig_customer="辽宁省公安厅")
        mapping = [{"current": "SF-1", "owner": "x", "closed": "SS-99"}]
        out = P.build_projects(ppm, {"佘海龙"}, set(), mapping, [], self.TOPMAP)
        sf = next(p for p in out if p["projectId"] == "SF-1")
        assert sf["isPresale"] is True
        assert sf["top1000"] == "是"
        assert sf["quadrant"] == "M1 战略核心区"

    def test_presale_ignores_own_customer_uses_original(self):
        # 本项目客户在清单(若用本项目会判是),但原项目客户不在清单 → 否(证明按原项目)
        ppm = self._ppm(own_customer="辽宁省公安厅", orig_customer="不在表里")
        mapping = [{"current": "SF-1", "owner": "x", "closed": "SS-99"}]
        out = P.build_projects(ppm, {"佘海龙"}, set(), mapping, [], self.TOPMAP)
        sf = next(p for p in out if p["projectId"] == "SF-1")
        assert sf["top1000"] == "否"
        assert sf["quadrant"] == ""

    def test_presale_no_mapping_is_no_even_if_own_in_list(self):
        # 售前无原项目映射:即便本项目客户在清单,也不回退本项目 → 否/空
        ppm = self._ppm(own_customer="辽宁省公安厅", orig_customer="x", orig_in_pm=False)
        out = P.build_projects(ppm, {"佘海龙"}, set(), [], [], self.TOPMAP)
        sf = next(p for p in out if p["projectId"] == "SF-1")
        assert sf["top1000"] == "否"
        assert sf["quadrant"] == ""

    def test_presale_original_not_in_pm_is_no(self):
        # 有映射但原项目不在 projectPmis → 取不到原客户 → 否/空
        ppm = self._ppm(own_customer="辽宁省公安厅", orig_customer="x", orig_in_pm=False)
        mapping = [{"current": "SF-1", "owner": "x", "closed": "SS-99"}]
        out = P.build_projects(ppm, {"佘海龙"}, set(), mapping, [], self.TOPMAP)
        sf = next(p for p in out if p["projectId"] == "SF-1")
        assert sf["top1000"] == "否"
        assert sf["quadrant"] == ""

    def test_non_presale_still_uses_own_customer(self):
        # 非售前:仍按本项目最终客户(回归保护)
        p = _pm_active("实施甲", "佘海龙")
        p["customer"]["最终客户"] = "辽宁省公安厅"
        out = P.build_projects({"SS-1": p}, {"佘海龙"}, set(), [], [], self.TOPMAP)
        assert out[0]["isPresale"] is False
        assert out[0]["top1000"] == "是"
        assert out[0]["quadrant"] == "M1 战略核心区"


def test_effective_sign_unit():
    from projects import effective_sign_unit
    assert effective_sign_unit(False, "本单位", "原单位") == "本单位"   # 非售前取本项目
    assert effective_sign_unit(True, "", "原单位") == "原单位"          # 售前本空→原项目
    assert effective_sign_unit(True, "本单位", "原单位") == "原单位"    # 售前恒取原项目(本值不覆盖)
    assert effective_sign_unit(False, "", "") == ""
    assert effective_sign_unit(True, "", "") == ""


class TestBuildProjectsSignUnit:
    """签约单位回退单一来源(Project.signUnit):非售前=本项目签约单位;售前=原项目签约单位(本项目该字段应为空)。"""

    def test_non_presale_uses_own_sign_unit(self):
        p = _pm_active("实施甲", "佘海龙")
        p["customer"]["签约单位"] = "本项目签约单位"
        out = P.build_projects({"SS-1": p}, {"佘海龙"}, set(), [], [])
        assert out[0]["signUnit"] == "本项目签约单位"

    def test_presale_falls_back_to_original_sign_unit(self):
        sf = _pm_active("售前甲", "佘海龙", project_type="售前服务类")
        sf["customer"]["签约单位"] = ""  # 售前本项目该字段应为空,以验证回退
        ss = {**_pm_active("原项目甲", "佘海龙"), "source": "已关闭"}
        ss["customer"]["签约单位"] = "原项目签约单位"
        ppm = {"SF-1": sf, "SS-99": ss}
        mapping = [{"current": "SF-1", "owner": "x", "closed": "SS-99"}]
        out = P.build_projects(ppm, {"佘海龙"}, set(), mapping, [])
        sf_out = next(p for p in out if p["projectId"] == "SF-1")
        assert sf_out["isPresale"] is True
        assert sf_out["signUnit"] == "原项目签约单位"
