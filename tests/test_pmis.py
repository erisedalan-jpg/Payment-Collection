# -*- coding: utf-8 -*-
"""pmis.py 纯函数单元测试。不依赖 input/ 真文件——用内存 dict 或 tmp_path 生成的小 xlsx。"""
import config
import openpyxl
import pytest
import pmis as M


class TestParsePmisMoney:
    def test_plain(self):
        assert M.parse_pmis_money("1234.5") == 1234.5
    def test_with_separators(self):
        assert M.parse_pmis_money("1,234,567") == 1234567.0
    def test_blank_is_none(self):
        assert M.parse_pmis_money("") is None
        assert M.parse_pmis_money(None) is None
    def test_number_passthrough(self):
        assert M.parse_pmis_money(1000) == 1000.0
    def test_negative(self):
        assert M.parse_pmis_money("-500.0") == -500.0


class TestParsePmisPct:
    def test_percent_text(self):
        assert M.parse_pmis_pct("80.00%") == pytest.approx(0.8)
    def test_bare_le_1(self):
        assert M.parse_pmis_pct(0.8) == pytest.approx(0.8)
    def test_gt_1_divided(self):
        assert M.parse_pmis_pct("100") == pytest.approx(1.0)
    def test_blank_none(self):
        assert M.parse_pmis_pct("") is None
        assert M.parse_pmis_pct(None) is None


class TestParseCloseFraction:
    """未关闭风险数量是 '未关闭/总' 分式文本,取分子。"""
    def test_fraction(self):
        assert M.parse_close_fraction("2/5") == 2
    def test_zero(self):
        assert M.parse_close_fraction("0/3") == 0
    def test_blank_none(self):
        assert M.parse_close_fraction("") is None
    def test_plain_int(self):
        assert M.parse_close_fraction("4") == 4


def _make_xlsx(dir_path, name, headers, rows):
    """造一个表头在第 2 行的 PMIS 风格 xlsx(第 1 行合并标题)。返回文件路径。"""
    import os as _os
    wb = openpyxl.Workbook(); ws = wb.active
    ws.cell(row=1, column=1, value="标题")
    for c, h in enumerate(headers, 1):
        ws.cell(row=2, column=c, value=h)
    for r, row in enumerate(rows, 3):
        for c, h in enumerate(headers, 1):
            ws.cell(row=r, column=c, value=row.get(h))
    _os.makedirs(dir_path, exist_ok=True)
    p = _os.path.join(dir_path, name); wb.save(p); return p


class TestReadPmisSheet:
    def test_reads_header_row2(self, tmp_path):
        p = _make_xlsx(str(tmp_path), "x.xlsx", ["项目编号", "项目名称"],
                       [{"项目编号": "A-1", "项目名称": "甲"}])
        rows = M.read_pmis_sheet(p)
        assert rows == [{"项目编号": "A-1", "项目名称": "甲"}]
    def test_missing_file_returns_empty(self, tmp_path):
        assert M.read_pmis_sheet(str(tmp_path / "nope.xlsx")) == []

    def test_reads_all_rows_not_truncated(self, tmp_path):
        # 回归:read_pmis_sheet 不得截断行(曾因 read_only 按不可靠 dimension 截断,真实 WPS 文件 911 行只读出 1 行)
        rows_in = [{"项目编号": f"P-{i}", "项目名称": f"项目{i}"} for i in range(30)]
        p = _make_xlsx(str(tmp_path), "many.xlsx", ["项目编号", "项目名称"], rows_in)
        out = M.read_pmis_sheet(p)
        assert len(out) == 30
        assert out[0]["项目编号"] == "P-0" and out[29]["项目编号"] == "P-29"


class TestDeriveCost:
    def test_consume_ratio_overrun_and_delivery(self):
        row = {"项目总预算（元）": "1000", "项目核算（元）": "600", "剩余预算（元）": "-50",
               "成本状态": "黄色预警"}
        center = {"是否交付部门人工成本超支": "是"}
        cost = M.derive_cost(row, center)
        assert cost["消耗比"] == pytest.approx(0.6)
        assert cost["项目超支"] is True       # 剩余预算 -50 < 0
        assert cost["交付超支"] is True        # 中心:是否交付部门人工成本超支==是
        assert cost["成本状态"] == "黄色预警"
        assert "超支" not in cost              # 旧键已移除

    def test_no_overrun(self):
        cost = M.derive_cost({"剩余预算（元）": "400"}, {"是否交付部门人工成本超支": "否"})
        assert cost["项目超支"] is False and cost["交付超支"] is False

    def test_zero_budget_ratio_none(self):
        cost = M.derive_cost({"项目总预算（元）": "0", "项目核算（元）": "0"}, {})
        assert cost["消耗比"] is None
        assert cost["交付超支"] is False


class TestDeriveRisk:
    def test_aggregate(self):
        recs = [{"风险等级": "低", "风险状态": "已关闭"},
                {"风险等级": "高", "风险状态": "已识别"}]
        risk = M.derive_risk(recs)
        assert risk["风险记录数"] == 2
        assert risk["最高等级"] == "高"
        assert risk["闭环率"] == pytest.approx(0.5)
    def test_empty(self):
        risk = M.derive_risk([])
        assert risk["风险记录数"] == 0 and risk["最高等级"] is None


class TestBuildProjectPmis:
    def _tables(self):
        active = {
            "base": [{"项目编号": "SS-1", "项目名称": "甲", "最终客户": "客A", "项目状态": "实施中"}],
            "center": [{"项目编号": "SS-1", "是否人工成本超支": "是"}],
            "status": [{"项目编号": "SS-1", "项目总预算（元）": "1000", "项目核算（元）": "500",
                        "项目累计完工进展百分比": "80%", "未关闭风险数量": "1/2"}],
            "risk": [{"项目编号": "SS-1", "风险等级": "高", "风险状态": "已识别"}],
        }
        closed = {
            "base": [{"项目编号": "SS-9", "项目名称": "乙", "项目状态": "已结项"},
                     {"项目编号": "SS-OUT", "项目名称": "丙"}],
            "center": [{"项目编号": "SS-9"}],
            "status": [{"项目编号": "SS-9", "项目总预算（元）": "200", "项目核算（元）": "200"}],
        }
        return active, closed

    def test_active_full_and_closed_filtered(self):
        active, closed = self._tables()
        pay_ids = {"SS-1", "SS-9", "SS-FREE"}  # SS-FREE 不在 PMIS;SS-OUT 在已关闭但不在回款
        pm = M.build_project_pmis(active, closed, pay_ids)
        assert "SS-1" in pm and pm["SS-1"]["matched"] is True
        assert pm["SS-1"]["source"] == "在建"
        assert pm["SS-1"]["cost"]["消耗比"] == pytest.approx(0.5)
        assert pm["SS-1"]["progress"]["完工进展"] == pytest.approx(0.8)
        assert pm["SS-1"]["risk"]["最高等级"] == "高"
        assert pm["SS-1"]["customer"]["最终客户"] == "客A"
        assert "SS-9" in pm and pm["SS-9"]["source"] == "已关闭"
        assert "SS-OUT" not in pm
        assert "SS-FREE" not in pm

    def test_pause_false_and_risk_override(self):
        active = {
            "base": [{"项目编号": "SS-2", "是否暂停": "否", "项目状态": "实施中"}],
            "center": [{"项目编号": "SS-2"}],
            "status": [{"项目编号": "SS-2", "未关闭风险数量": "3/5"}],
            "risk": [{"项目编号": "SS-2", "风险等级": "低", "风险状态": "已识别"}],
        }
        pm = M.build_project_pmis(active, {}, set())
        # "否" 必须判为 False(而非 None,也不被 "不是" 之类子串误判)
        assert pm["SS-2"]["status"]["是否暂停"] is False
        # status 表分式分子(3)覆盖 risk 记录推导值(1)
        assert pm["SS-2"]["risk"]["未关闭风险数"] == 3

    def test_active_universe_is_center_only(self):
        active = {
            "base": [{"项目编号": "ONLY-BASE", "项目名称": "x"},
                     {"项目编号": "IN-CENTER", "项目名称": "c"}],
            "center": [{"项目编号": "IN-CENTER"}],
            "status": [{"项目编号": "ONLY-STATUS"}],
            "risk": [],
        }
        pm = M.build_project_pmis(active, {}, set())
        assert "IN-CENTER" in pm
        assert "ONLY-BASE" not in pm and "ONLY-STATUS" not in pm


class TestComputeDataQuality:
    def test_unmatched_and_summary(self):
        project_pmis = {"SS-1": {"matched": True, "source": "在建",
                                 "cost": {"成本状态": None, "消耗比": 0.5},
                                 "progress": {"完工进展": None}, "status": {"项目状态": "实施中"}}}
        pay_projects = [
            {"projectId": "SS-1", "projectName": "甲"},
            {"projectId": "SS-9", "projectName": "乙"},
            {"projectId": "SF-2", "projectName": "丙售前"},
        ]
        project_pmis["SS-9"] = {"matched": True, "source": "已关闭",
                                "cost": {"成本状态": "正常", "消耗比": 1.0},
                                "progress": {"完工进展": 1.0}, "status": {"项目状态": "已结项"}}
        dq = M.compute_data_quality(project_pmis, pay_projects)
        assert dq["summary"]["matchedActive"] == 1
        assert dq["summary"]["matchedClosed"] == 1
        assert dq["summary"]["unmatched"] == 1
        kinds = {u["projectId"]: u["kind"] for u in dq["unmatched"]}
        assert kinds == {"SF-2": "SF售前"}
        assert dq["summary"]["pmisProvided"] is True
        assert dq["summary"]["joinRate"] == round(2 / 3, 4)
        assert dq["conflicts"] == M.PMIS_CONFLICTS
        assert dq["dirty"] == []
        bf = {b["projectId"]: b["missingFields"] for b in dq["backfill"]}
        assert "SS-1" in bf
        assert "完工进展" in bf["SS-1"] and "成本状态" in bf["SS-1"]


class TestPmisDataTime:
    def test_empty_dir_returns_blank(self, tmp_path):
        assert M.pmis_data_time(str(tmp_path)) == ''
    def test_missing_dir_returns_blank(self, tmp_path):
        assert M.pmis_data_time(str(tmp_path / 'nope')) == ''
    def test_returns_formatted_max_mtime(self, tmp_path):
        d = tmp_path / "pmis"
        d.mkdir()
        (d / "项目中心.xlsx").write_bytes(b"x")
        out = M.pmis_data_time(str(d))
        assert len(out) == 16 and out[4] == '-' and out[13] == ':'


class TestLoadProjectPmis:
    def test_missing_dir_graceful(self, tmp_path):
        pm, dq = M.load_project_pmis(str(tmp_path / "nope"), {"SS-1"})
        assert pm == {}
        assert dq["summary"]["pmisProvided"] is False
        assert dq["summary"]["lastPmisUpdate"] == ''

    def test_reads_files(self, tmp_path):
        d = tmp_path / "pmis"
        d.mkdir()
        _make_xlsx(str(d), config.PMIS_FILES_ACTIVE["base"],
                   ["项目编号", "项目名称", "项目状态"],
                   [{"项目编号": "SS-1", "项目名称": "甲", "项目状态": "实施中"}])
        _make_xlsx(str(d), config.PMIS_FILES_ACTIVE["center"],
                   ["项目编号"],
                   [{"项目编号": "SS-1"}])
        _make_xlsx(str(d), config.PMIS_FILES_ACTIVE["status"],
                   ["项目编号", "项目总预算（元）", "项目核算（元）"],
                   [{"项目编号": "SS-1", "项目总预算（元）": "1000", "项目核算（元）": "500"}])
        pm, dq = M.load_project_pmis(str(d), {"SS-1"})
        assert "SS-1" in pm
        assert dq["summary"]["pmisProvided"] is True
        assert len(dq["summary"]["lastPmisUpdate"]) == 16

    def test_empty_sheets_graceful(self, tmp_path):
        d = tmp_path / "pmis"
        d.mkdir()
        # 目录存在但无任何 xlsx → read_pmis_sheet 全返回 [] → 优雅降级
        pm, dq = M.load_project_pmis(str(d), {"SS-1"})
        assert pm == {}
        assert dq["summary"]["pmisProvided"] is False


class TestAssembleTeamAndRisks:
    def test_team_from_center_then_base(self):
        base_i = {"P1": {"项目经理（FR）": "李四", "项目经理L4部门": "银行服务组", "项目名称": "B名"}}
        center_i = {"P1": {"项目经理": "张三", "项目名称": "C名"}}
        out = M._assemble("P1", base_i, center_i, {}, {}, "在建")
        assert out["team"]["项目名称"] == "C名"
        assert out["team"]["项目经理"] == "张三"
        assert out["team"]["L4部门"] == "银行服务组"
        assert out["team"]["AR"] is None        # base 无该列 → None

    def test_team_extended_fields_from_base(self):
        base_i = {"P1": {"项目经理L3部门": "三部", "项目经理L3-1部门": "三部一组",
                         "客户经理（AR）": "AR人", "方案经理（SR）": "SR人",
                         "安全运行经理（CSR）": "CSR人", "定制经理（CDR）": "CDR人",
                         "Sponsor": "老板"}}
        t = M._assemble("P1", base_i, {}, {}, {}, "在建")["team"]
        assert t["L3部门"] == "三部" and t["L3_1部门"] == "三部一组"
        assert t["AR"] == "AR人" and t["SR"] == "SR人" and t["CSR"] == "CSR人"
        assert t["CDR"] == "CDR人" and t["Sponsor"] == "老板"
        assert set(t.keys()) == {"项目名称", "项目经理", "L4部门", "L3部门", "L3_1部门", "AR", "SR", "CSR", "CDR", "Sponsor"}

    def test_customer_signing_unit_and_contract_center_priority(self):
        base_i = {"P1": {"签约单位": "甲方单位", "合同编号": "B-001", "最终客户": "客A",
                         "行业中类": "金融", "合同总额（元）": "1000"}}
        center_i = {"P1": {"合同编号": "C-001"}}
        cust = M._assemble("P1", base_i, center_i, {}, {}, "在建")["customer"]
        assert cust["签约单位"] == "甲方单位"
        assert cust["合同编号"] == "C-001"      # center 优先
        assert "签约形式" not in cust
        cust2 = M._assemble("P1", base_i, {}, {}, {}, "在建")["customer"]
        assert cust2["合同编号"] == "B-001"      # center 缺 → 回退 base

    def test_status_key_action_and_deliverable(self):
        status_i = {"P1": {"关键动作完成情况(必须-考核)": "已完成",
                           "交付物上传情况(必须-考核)": "3/3"}}
        st = M._assemble("P1", {}, {}, status_i, {}, "在建")["status"]
        assert st["关键动作"] == "已完成" and st["交付物"] == "3/3"

    def test_team_fallback_to_base(self):
        base_i = {"P1": {"项目经理（FR）": "李四", "项目经理L4部门": "银行服务组", "项目名称": "B名"}}
        out = M._assemble("P1", base_i, {}, {}, {}, "在建")
        assert out["team"]["项目经理"] == "李四"
        assert out["team"]["项目名称"] == "B名"

    def test_status_level_and_type_from_base_then_status(self):
        # 项目级别/项目类型:base 优先 status 兜底(与"项目状态"同模式)
        base_i = {"P1": {"项目级别": "P3", "项目类型": "交付项目"}}
        out = M._assemble("P1", base_i, {}, {}, {}, "在建")
        assert out["status"]["项目级别"] == "P3"
        assert out["status"]["项目类型"] == "交付项目"
        # base 缺 → status 兜底
        status_i = {"P2": {"项目级别": "P1", "项目类型": "实施项目"}}
        out2 = M._assemble("P2", {}, {}, status_i, {}, "在建")
        assert out2["status"]["项目级别"] == "P1"
        assert out2["status"]["项目类型"] == "实施项目"
        # 两表皆缺 → None
        out3 = M._assemble("P3", {}, {}, {}, {}, "在建")
        assert out3["status"]["项目级别"] is None and out3["status"]["项目类型"] is None

    def test_risk_records_jsonable(self):
        import datetime
        risk_i = {"P1": [{"风险等级": "高", "风险状态": "已关闭",
                          "登记日期": datetime.datetime(2026, 1, 2, 3, 4)}]}
        out = M._assemble("P1", {}, {}, {}, risk_i, "在建")
        recs = out["riskRecords"]
        assert len(recs) == 1
        assert recs[0]["登记日期"] == "2026-01-02T03:04:00"  # datetime 必须转 str 才能入 JSON

    def test_risk_records_timedelta_safe(self):
        import datetime
        risk_i = {"P1": [{"耗时": datetime.timedelta(hours=1, minutes=30)}]}
        out = M._assemble("P1", {}, {}, {}, risk_i, "在建")
        v = out["riskRecords"][0]["耗时"]
        assert isinstance(v, str)  # 任何非 JSON 原生类型都必须转 str


class TestBuildProjectPmisExtraClosed:
    def test_closed_included_via_extra_ids(self):
        closed = {"base": [{"项目编号": "SS-1", "项目状态": "已完工"}], "center": [], "status": []}
        out = M.build_project_pmis({"base": [], "center": [], "status": [], "risk": []},
                                   closed, set(), extra_closed_ids={"SS-1"})
        assert "SS-1" in out and out["SS-1"]["source"] == "已关闭"

    def test_closed_excluded_without_any_ids(self):
        closed = {"base": [{"项目编号": "SS-1"}], "center": [], "status": []}
        out = M.build_project_pmis({"base": [], "center": [], "status": [], "risk": []},
                                   closed, set())
        assert "SS-1" not in out


class TestBuildClosedProjects:
    def _make_closed(self, tmp_path):
        import openpyxl, datetime
        d = tmp_path / "pmis"; d.mkdir()

        def _wb(fn, headers, rows):
            wb = openpyxl.Workbook(); ws = wb.active
            ws.cell(row=1, column=1, value="标题")  # 第1行为合并标题,表头在第2行
            for j, h in enumerate(headers, start=1):
                ws.cell(row=2, column=j, value=h)
            for i, rec in enumerate(rows, start=3):
                for j, h in enumerate(headers, start=1):
                    ws.cell(row=i, column=j, value=rec.get(h))
            wb.save(str(d / fn))

        _wb(M.config.PMIS_FILES_CLOSED["center"],
            ["项目编号", "项目名称", "项目经理", "是否交付部门人工成本超支", "成本状态", "计划终验时间", "合同编号"],
            [{"项目编号": "C-1", "项目名称": "中心甲", "项目经理": "张三",
              "是否交付部门人工成本超支": "是", "成本状态": "正常",
              "计划终验时间": datetime.datetime(2025, 7, 1), "合同编号": "HT-C1"},
             {"项目编号": "C-2", "项目经理": "外部人"}])
        _wb(M.config.PMIS_FILES_CLOSED["base"],
            ["项目编号", "项目名称", "项目经理（FR）", "项目经理L4部门", "项目经理L3-1部门",
             "签约单位", "最终客户", "行业中类", "合同总额（元）", "合同编号",
             "项目状态", "项目关闭时间", "是否正常关闭", "关闭说明"],
            [{"项目编号": "C-1", "项目名称": "基础甲", "项目经理L4部门": "安全A组",
              "项目经理L3-1部门": "三部一组", "签约单位": "甲单位", "最终客户": "客A",
              "行业中类": "金融", "合同总额（元）": "1000000", "合同编号": "HT-B1",
              "项目状态": "已验收", "项目关闭时间": datetime.datetime(2025, 8, 15),
              "是否正常关闭": "是", "关闭说明": "正常结项"}])
        _wb(M.config.PMIS_FILES_CLOSED["status"],
            ["项目编号", "项目总预算（元）", "项目核算（元）", "剩余预算（元）",
             "项目阶段", "项目类型", "项目级别", "项目累计完工进展百分比"],
            [{"项目编号": "C-1", "项目总预算（元）": "1000", "项目核算（元）": "1200",
              "剩余预算（元）": "-200", "项目阶段": "项目收尾", "项目类型": "实施项目",
              "项目级别": "B", "项目累计完工进展百分比": "100"}])
        return str(d)

    def test_universe_and_fields(self, tmp_path):
        d = self._make_closed(tmp_path)
        out = M.build_closed_projects(d, {"张三"})
        assert [p["projectId"] for p in out] == ["C-1"]      # 仅经理∈org_names(C-2 外部人剔除)
        p = out[0]
        assert p["projectName"] == "中心甲"                   # center 优先
        assert p["projectManager"] == "张三"
        assert p["orgL4"] == "安全A组" and p["orgL3_1"] == "三部一组"
        assert p["合同编号"] == "HT-C1"                        # center 优先→base
        assert p["customer"]["签约单位"] == "甲单位"
        assert p["customer"]["合同总额"] == 1000000.0
        assert p["status"]["项目状态"] == "已验收" and p["status"]["项目级别"] == "B"
        assert p["progress"]["项目阶段"] == "项目收尾"
        assert p["cost"]["剩余预算"] == -200.0 and p["cost"]["项目超支"] is True
        assert p["cost"]["交付超支"] is True                   # center 是否交付部门人工成本超支==是
        assert p["closeInfo"]["关闭时间"] == "2025-08-15"      # datetime→YYYY-MM-DD
        assert p["closeInfo"]["计划终验时间"] == "2025-07-01"
        assert p["closeInfo"]["是否正常关闭"] == "是"
        assert p["team"]["L3_1部门"] == "三部一组"             # 下划线键

    def test_empty_org_no_filter_and_missing_dir(self, tmp_path):
        d = self._make_closed(tmp_path)
        assert {p["projectId"] for p in M.build_closed_projects(d, set())} == {"C-1", "C-2"}  # 空清单不过滤
        assert M.build_closed_projects(str(tmp_path / "none"), {"张三"}) == []                 # 缺目录→[]

    def test_count_consistency(self, tmp_path):
        import projects as P
        d = self._make_closed(tmp_path)
        assert len(M.build_closed_projects(d, {"张三"})) == P.count_closed_dept(d, {"张三"})


class TestPmisDateToStr:
    def test_datetime(self):
        import datetime
        assert M._pmis_date_to_str(datetime.datetime(2025, 8, 15)) == "2025-08-15"
        assert M._pmis_date_to_str(datetime.date(2025, 7, 1)) == "2025-07-01"

    def test_string_truncates_to_date(self):
        assert M._pmis_date_to_str("2025-08-15 00:00:00") == "2025-08-15"
        assert M._pmis_date_to_str("  2025-08-15  ") == "2025-08-15"

    def test_none_and_empty(self):
        assert M._pmis_date_to_str(None) is None
        assert M._pmis_date_to_str("") is None
