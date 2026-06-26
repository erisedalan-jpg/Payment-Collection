import openpyxl
import opportunities as opp


def _store():
    return {"version": 1, "seq": 0, "rows": []}


def test_new_row_blank_has_all_fields():
    r = opp.new_row("opp-1")
    assert r["id"] == "opp-1"
    for f in opp.FIELDS:
        assert f in r and r[f] == "" or f == "amountWan"
    assert r["amountWan"] == "" and r["firstReg"] == "" and r["lastUpdate"] == ""


def test_apply_create_increments_seq():
    s = _store()
    a = opp.apply_create(s, "2026-06-24")
    b = opp.apply_create(s, "2026-06-24")
    assert s["seq"] == 2 and a["id"] == "opp-1" and b["id"] == "opp-2"
    assert len(s["rows"]) == 2


def test_apply_update_stamps_firstreg_only_when_content_then_lastupdate_each_time():
    s = _store()
    opp.apply_create(s, "2026-06-24")  # opp-1, 空
    # 首次写入有内容 → firstReg 盖
    r = opp.apply_update(s, "opp-1", {"customer": "甲公司"}, "admin", "2026-06-24", "2026-06-24 10:00")
    assert r["customer"] == "甲公司"
    assert r["firstReg"] == "2026-06-24"
    assert r["lastUpdate"] == "2026-06-24 10:00" and r["lastUpdateBy"] == "admin"
    # 二次更新 → firstReg 不变, lastUpdate 刷新
    r2 = opp.apply_update(s, "opp-1", {"status": "招投标"}, "admin", "2026-06-25", "2026-06-25 09:00")
    assert r2["firstReg"] == "2026-06-24" and r2["lastUpdate"] == "2026-06-25 09:00"


def test_apply_update_firstreg_not_set_when_all_blank():
    s = _store(); opp.apply_create(s, "2026-06-24")
    r = opp.apply_update(s, "opp-1", {"customer": ""}, "admin", "2026-06-24", "2026-06-24 10:00")
    assert r["firstReg"] == ""  # 无内容不盖首登
    assert r["lastUpdate"] == "2026-06-24 10:00"


def test_apply_update_rejects_unknown_field_and_missing_row():
    s = _store(); opp.apply_create(s, "2026-06-24")
    r = opp.apply_update(s, "opp-1", {"evil": "x", "id": "hack"}, "admin", "d", "t")
    assert "evil" not in r and r["id"] == "opp-1"  # 非 FIELDS 被拒
    assert opp.apply_update(s, "nope", {"customer": "x"}, "admin", "d", "t") is None


def test_apply_update_parses_amount_and_dates():
    s = _store(); opp.apply_create(s, "2026-06-24")
    r = opp.apply_update(s, "opp-1", {"amountWan": "1,200.5", "bidDate": "2026-07-01 00:00:00"}, "a", "d", "t")
    assert r["amountWan"] == 1200.5 and r["bidDate"] == "2026-07-01"


def test_apply_delete():
    s = _store(); opp.apply_create(s, "d"); opp.apply_create(s, "d")
    assert opp.apply_delete(s, ["opp-1"]) == 1
    assert [r["id"] for r in s["rows"]] == ["opp-2"]


def test_filter_for_account():
    rows = [{"id": "1", "l4": "小金融服务组"}, {"id": "2", "l4": "银行服务组"}]
    assert len(opp.filter_for_account(rows, [], True)) == 2          # 超管全看
    assert len(opp.filter_for_account(rows, ["*"], False)) == 2       # '*' 全看
    assert [r["id"] for r in opp.filter_for_account(rows, ["小金融服务组"], False)] == ["1"]
    assert opp.filter_for_account(rows, [], False) == []             # 空 allowedL4 → 无


def test_read_xlsx_maps_headers(tmp_path):
    p = tmp_path / "opportunities.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["客户名称", "L4组织", "商机状态", "预估金额（万元）"])
    ws.append(["甲公司", "小金融服务组", "招投标", 320])
    wb.save(p)
    rows = opp.read_opportunities_xlsx(str(p))
    assert len(rows) == 1
    assert rows[0]["customer"] == "甲公司" and rows[0]["l4"] == "小金融服务组"
    assert rows[0]["status"] == "招投标" and rows[0]["amountWan"] == 320.0
    assert rows[0]["id"] == "opp-1"


def test_read_xlsx_missing_file():
    assert opp.read_opportunities_xlsx("nonexistent.xlsx") == []


def test_read_xlsx_halfwidth_amount_header(tmp_path):
    """M3: 前端导出用半角括号表头 '预估金额(万元)'，导入时 amountWan 须被正确解析（非空）。"""
    p = tmp_path / "opp_halfwidth.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["客户名称", "L4组织", "商机状态", "预估金额(万元)"])  # 半角括号
    ws.append(["乙公司", "大金融服务组", "需求挖掘", 580])
    wb.save(p)
    rows = opp.read_opportunities_xlsx(str(p))
    assert len(rows) == 1
    assert rows[0]["amountWan"] == 580.0, (
        f"半角括号表头的金额列被静默丢弃，amountWan={rows[0]['amountWan']!r}"
    )


def test_apply_create_with_fields_stamps_content():
    s = _store()
    r = opp.apply_create_with_fields(s, {"customer": "丙公司", "amountWan": "88"},
                                     "admin", "2026-06-25", "2026-06-25 12:00")
    assert r["id"] == "opp-1" and r["customer"] == "丙公司" and r["amountWan"] == 88.0
    assert r["firstReg"] == "2026-06-25"          # 有内容 → 盖首登
    assert r["lastUpdate"] == "2026-06-25 12:00" and r["lastUpdateBy"] == "admin"
    assert len(s["rows"]) == 1


def test_apply_create_with_fields_empty_is_blank_row():
    s = _store()
    r = opp.apply_create_with_fields(s, None, "admin", "2026-06-25", "2026-06-25 12:00")
    assert r["id"] == "opp-1" and r["customer"] == "" and r["firstReg"] == ""  # 无内容不盖首登
    assert len(s["rows"]) == 1


def test_opportunity_level_is_editable_field():
    assert 'opportunityLevel' in opp.FIELDS
    r = opp.new_row("opp-1")
    assert r["opportunityLevel"] == ""
    s = _store(); opp.apply_create(s, "d")
    r2 = opp.apply_update(s, "opp-1", {"opportunityLevel": "P2"}, "admin", "d", "t")
    assert r2["opportunityLevel"] == "P2"


def test_read_xlsx_maps_opportunity_level(tmp_path):
    p = tmp_path / "opp_level.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["客户名称", "商机级别"])
    ws.append(["甲公司", "P1"])
    wb.save(p)
    rows = opp.read_opportunities_xlsx(str(p))
    assert rows[0]["opportunityLevel"] == "P1"


def test_major_poc_is_editable_field():
    assert 'majorPoc' in opp.FIELDS
    r = opp.new_row("opp-1")
    assert r["majorPoc"] == ""
    s = _store(); opp.apply_create(s, "d")
    r2 = opp.apply_update(s, "opp-1", {"majorPoc": "是"}, "admin", "d", "t")
    assert r2["majorPoc"] == "是"


def test_read_xlsx_maps_major_poc(tmp_path):
    p = tmp_path / "opp_poc.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["客户名称", "是否重大POC"])
    ws.append(["甲公司", "是"])
    wb.save(p)
    rows = opp.read_opportunities_xlsx(str(p))
    assert rows[0]["majorPoc"] == "是"


def test_can_access_l4_super_and_wildcard_always_true():
    # 超管恒 True（即便 allowedL4 为空）
    assert opp.can_access_l4("银行服务组", [], True) is True
    assert opp.can_access_l4("", [], True) is True
    # allowedL4 含 '*' → 恒 True
    assert opp.can_access_l4("任意组", ["*"], False) is True


def test_can_access_l4_normal_admin_scoped():
    allowed = ["小金融服务组", "银行服务组"]
    # 命中本人 L4 → True
    assert opp.can_access_l4("小金融服务组", allowed, False) is True
    assert opp.can_access_l4("银行服务组", allowed, False) is True
    # 越权 L4 → False（防普通管理员改/建本人范围外的商机）
    assert opp.can_access_l4("运营商服务组", allowed, False) is False
    # 空 L4 值 / 空 allowedL4 → False
    assert opp.can_access_l4("", allowed, False) is False
    assert opp.can_access_l4("小金融服务组", [], False) is False
