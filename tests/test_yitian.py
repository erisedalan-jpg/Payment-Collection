# -*- coding: utf-8 -*-
"""yitian.py 管线单测。不依赖真 input/——xlsx/csv 全部用 tmp_path 现造。"""
import os

import openpyxl
import pytest

import projects as P
import yitian as Y

TS_HEADERS = [
    "ID", "工时类型", "客户", "项目类型", "工作类型三", "产研侧产品线", "产研侧产品名称",
    "工作日", "工时", "销售L2组织", "员工编号", "员工", "员工电话", "L4组织",
    "工作成果", "工单编号", "服务方式", "工时ID",
]
GOOD = "工作概述:巡检。工作进展:已完成。下一步工作计划:回访。"


def _ts_row(**kw):
    base = {
        "ID": "1", "工时类型": "项目类", "客户": "某客户", "项目类型": "交付实施",
        "工作类型三": "安装部署", "产研侧产品线": "", "产研侧产品名称": "",
        "工作日": "2026-06-01", "工时": 8, "销售L2组织": "银行集团军",
        "员工编号": "a012804", "员工": "佘海龙", "员工电话": "13500000000",
        "L4组织": "工时表里的脏组织", "工作成果": GOOD, "工单编号": "WO1", "服务方式": "远程",
    }
    base.update(kw)
    if "工时ID" not in kw:
        base["工时ID"] = base["ID"]              # 默认工时ID与ID同值(未显式传时随ID一起变)
    return [base[h] for h in TS_HEADERS]


def _write_timesheet(base_dir, ts_rows, ts_headers=None, drop_cols=None):
    """只(重)写 input/yitian/工时.xlsx,不动组织架构/TOP1000——模拟"第二周只含当周数据
    的新导出"覆盖旧文件。ts_headers 可覆盖表头,drop_cols 从(默认)表头里摘掉指定列
    (缺列场景用,行值随之同步裁剪保持位置对齐)。"""
    ydir = os.path.join(str(base_dir), "input", "yitian")
    os.makedirs(ydir, exist_ok=True)
    headers = ts_headers if ts_headers is not None else TS_HEADERS
    if drop_cols:
        headers = [h for h in headers if h not in drop_cols]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for r in ts_rows:
        row_map = dict(zip(TS_HEADERS, r))
        ws.append([row_map[h] for h in headers])
    wb.save(os.path.join(ydir, "工时.xlsx"))


def _make_input(tmp_path, ts_rows, org_rows=None, top_rows=None, holidays=None, ts_headers=None,
                 drop_cols=None):
    """造 input/ 目录树:input/yitian/工时.xlsx + input/组织架构.xlsx + input/TOP1000.xlsx。
    ts_headers 可覆盖表头、drop_cols 摘掉指定列(两者都是缺列场景用),默认 TS_HEADERS 全量。"""
    base = tmp_path
    ydir = base / "input" / "yitian"
    ydir.mkdir(parents=True)

    _write_timesheet(str(base), ts_rows, ts_headers=ts_headers, drop_cols=drop_cols)

    org = openpyxl.Workbook()
    ows = org.active
    ows.append(["工号", "姓名", "员工类别", "新L2组织", "新L3组织", "新L3-1组织", "新L4组织"])
    for r in (org_rows if org_rows is not None else
              [("A012804", "佘海龙", "正式员工", "交付中心", "交付实施三部", "服务二部", "银行服务组")]):
        ows.append(list(r))
    org.save(str(base / "input" / "组织架构.xlsx"))

    top = openpyxl.Workbook()
    tws = top.active
    tws.append(["客户编号", "客户名称", "客户级别"])
    for r in (top_rows if top_rows is not None else [("C1", "某客户", "TOP1000大客户")]):
        tws.append(list(r))
    top.save(str(base / "input" / "TOP1000.xlsx"))

    if holidays:
        (ydir / "holidays.csv").write_text(
            "\n".join(["日期,类型"] + [f"{d},{k}" for d, k in holidays]), encoding="utf-8")
    return str(base)


class TestReadOrgRoster:
    def test_upper_normalizes_and_filters_dept(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["工号", "姓名", "员工类别", "新L2组织", "新L3组织", "新L3-1组织", "新L4组织"])
        ws.append(["a012804", "佘海龙", "正式员工", "交付中心", "交付实施三部", "服务二部", "银行服务组"])
        ws.append(["B000001", "外部门", "正式员工", "交付中心", "交付实施一部", "别的部", "别的组"])
        ws.append(["", "无工号", "正式员工", "交付中心", "交付实施三部", "服务二部", "银行服务组"])
        path = str(tmp_path / "组织架构.xlsx")
        wb.save(path)

        roster = P.read_org_roster(path)
        assert [p["id"] for p in roster] == ["A012804"]        # 大写归一 + 只留三部 + 丢无工号
        assert roster[0]["l4"] == "银行服务组"
        assert roster[0]["l31"] == "服务二部"


class TestBuildYitianData:
    def test_missing_timesheet_returns_none(self, tmp_path):
        # 现在数据来自累积库:input/yitian/工时.xlsx 缺失 → ingest() 不动累积库,
        # 累积库仍为空 → build_yitian_data() 也 None。
        (tmp_path / "input").mkdir()
        assert Y.ingest(str(tmp_path)) is None
        assert Y.build_yitian_data(str(tmp_path)) is None

    def test_missing_required_column_returns_none_and_logs_error(self, tmp_path, capsys):
        # I-3: 白名单列一旦被导出端改名/删列,不能静默错判(全量误判 MISS_SERVICE_MODE 却零报错)。
        # 缺列校验现在发生在 ingest()(读 xlsx 的唯一入口);ingest 失败不写累积库,
        # build_yitian_data() 因累积库仍空而返回 None。
        base = _make_input(tmp_path, [_ts_row()], drop_cols=["服务方式"])

        assert Y.ingest(base) is None
        out = capsys.readouterr().out
        assert "[ERROR]" in out
        assert "服务方式" in out
        assert Y.build_yitian_data(base) is None

    def test_basic_shape(self, tmp_path):
        base = _make_input(tmp_path, [_ts_row()])
        Y.ingest(base)
        data = Y.build_yitian_data(base)
        assert data["meta"]["rows"] == 1
        assert data["meta"]["employees"] == 1
        assert data["meta"]["periodStart"] == "2026-06-01"
        assert data["meta"]["calendarSource"] == "fallback"    # 没给 holidays.csv
        assert data["meta"]["hoursPerDay"] == 8
        assert "交付中心" in data["meta"]["thisBgL2"]
        e = data["entries"][0]
        assert e["e"] == "A012804"                              # 工号大写归一
        assert e["h"] == 8
        assert e["ok"] == 0 and e["iss"] == []
        assert e["top"] is True                                 # 客户命中 TOP1000
        assert data["dims"]["types"][e["t"]] == "项目类"
        assert data["issues"] == []

    def test_privacy_no_phone_and_no_content_for_clean_rows(self, tmp_path):
        base = _make_input(tmp_path, [_ts_row()])
        Y.ingest(base)
        data = Y.build_yitian_data(base)
        blob = repr(data)
        assert "13500000000" not in blob                        # 电话绝不落盘
        assert GOOD not in blob                                 # 合规行不下发工作成果正文

    def test_issue_row_gets_snippet(self, tmp_path):
        base = _make_input(tmp_path, [_ts_row(工作成果="今天干了点活", 服务方式="")])
        Y.ingest(base)
        data = Y.build_yitian_data(base)
        e = data["entries"][0]
        assert e["ok"] == 2
        assert "MISS_SUMMARY" in e["iss"] and "MISS_SERVICE_MODE" in e["iss"]
        iss = data["issues"][0]
        assert iss["i"] == 0
        assert iss["snippet"] == "今天干了点活"
        assert len(iss["codes"]) == len(iss["msgs"])

    def test_hint_only_row_gets_no_snippet(self, tmp_path):
        # I-6(用户裁决:收紧):只有真问题行(ok=2)才下发 120 字摘要;
        # 合规(提示)行(ok=1,如 HINT_PRESALE_PRODUCT)不下发工作成果正文。
        content = GOOD + "这是提示行的工作成果全文不应下发"
        base = _make_input(tmp_path, [_ts_row(
            项目类型="售前服务类", 工作类型三="环境调研", 产研侧产品线="其他",
            工作成果=content,
        )])
        Y.ingest(base)
        data = Y.build_yitian_data(base)
        e = data["entries"][0]
        assert e["ok"] == 1
        assert e["iss"] == ["HINT_PRESALE_PRODUCT"]
        iss = data["issues"][0]
        assert iss["codes"] == ["HINT_PRESALE_PRODUCT"]
        assert iss["snippet"] == ""                              # 提示行 snippet 为空串
        assert content not in repr(data)                          # 提示行正文绝不落盘

    def test_excluded_type_still_gets_entry_but_no_codes(self, tmp_path):
        # 假期类没有必填字段规则 → 判定结果为空码/ok=0,但仍是一条正常 entry;
        # 是否计入合规率由前端按超管配置的 excludedTypes 决定,后端不再预判(不再有 chk 字段)。
        base = _make_input(tmp_path, [_ts_row(工时类型="假期类", 工作成果="", 客户="", 服务方式="")])
        Y.ingest(base)
        data = Y.build_yitian_data(base)
        e = data["entries"][0]
        assert "chk" not in e
        assert e["ok"] == 0 and e["iss"] == []
        assert data["issues"] == []

    def test_zero_hour_row_is_still_checked(self, tmp_path):
        # 0 工时行现在照常检查(与原工具一致;原脚本 README 声称跳过、代码并没跳过)
        base = _make_input(tmp_path, [_ts_row(工时=0, 工作成果="今天干了点活", 服务方式="")])
        Y.ingest(base)
        data = Y.build_yitian_data(base)
        e = data["entries"][0]
        assert e["h"] == 0
        assert e["ok"] == 2                       # 正文缺三段 → 判问题
        assert "MISS_SUMMARY" in e["iss"]

    def test_org_columns_come_from_roster_not_timesheet(self, tmp_path):
        base = _make_input(tmp_path, [_ts_row()])
        Y.ingest(base)
        data = Y.build_yitian_data(base)
        assert data["roster"][0]["l4"] == "银行服务组"           # 不是工时表里的"工时表里的脏组织"
        assert "工时表里的脏组织" not in repr(data)

    def test_unknown_employee_dropped_and_counted(self, tmp_path):
        # 两行工时ID须不同(累积库按 wid 去重,同 wid 会互相覆盖导致行数错判)
        base = _make_input(tmp_path, [
            _ts_row(ID="1"), _ts_row(ID="2", 员工编号="Z999999", 员工="离职的"),
        ])
        Y.ingest(base)
        data = Y.build_yitian_data(base)
        assert data["meta"]["rows"] == 1
        assert data["meta"]["droppedRows"] == 1

    def test_presale_service_corrected_to_project_type(self, tmp_path):
        base = _make_input(tmp_path, [_ts_row(工时类型="售前类", 项目类型="售前服务类")])
        Y.ingest(base)
        data = Y.build_yitian_data(base)
        e = data["entries"][0]
        assert data["dims"]["types"][e["t"]] == "项目类"

    def test_holidays_csv_switches_source_and_days(self, tmp_path):
        base = _make_input(
            tmp_path,
            [_ts_row(工作日="2026-06-01"), _ts_row(ID="2", 工作日="2026-06-03")],
            holidays=[("2026-06-02", "休")],
        )
        Y.ingest(base)
        data = Y.build_yitian_data(base)
        assert data["meta"]["calendarSource"] == "csv"
        by_d = {d["d"]: d["workday"] for d in data["days"]}
        assert by_d["2026-06-01"] is True
        assert by_d["2026-06-02"] is False                      # 法定假(本是周二)
        assert by_d["2026-06-03"] is True

    def test_same_workorder_peer_rescues_product(self, tmp_path):
        # 两条同工单:A 只写防火墙(他家词) / B 写了 SOAR(本产品词) → A 应被同工单关联救回
        base = _make_input(tmp_path, [
            _ts_row(ID="1", 产研侧产品线="NGSOC", 工作成果=GOOD + "更换防火墙策略", 工单编号="WO9"),
            _ts_row(ID="2", 产研侧产品线="NGSOC", 工作成果=GOOD + "处理SOAR告警", 工单编号="WO9"),
        ])
        Y.ingest(base)
        data = Y.build_yitian_data(base)
        assert all("PRODUCT_MISMATCH" not in e["iss"] for e in data["entries"])

    def test_empty_timesheet_yields_empty_store_and_none(self, tmp_path):
        # 工时表 0 数据行 → ingest 后累积库仍是空的(没有可 upsert 的行)→
        # build_yitian_data() 按"累积库为空"语义返回 None(不是"空壳 dict")。
        base = _make_input(tmp_path, [])
        r = Y.ingest(base)
        assert r == {"added": 0, "updated": 0, "total": 0}
        assert Y.build_yitian_data(base) is None


class TestIngestAndAccumulate:
    def test_ingest_missing_file_returns_none(self, tmp_path):
        (tmp_path / "input" / "yitian").mkdir(parents=True)
        assert Y.ingest(str(tmp_path)) is None

    def test_ingest_then_build(self, tmp_path):
        base = _make_input(tmp_path, [_ts_row()])
        r = Y.ingest(base)
        assert r == {"added": 1, "updated": 1 - 1, "total": 1}
        data = Y.build_yitian_data(base)
        assert data["meta"]["rows"] == 1

    def test_second_week_accumulates(self, tmp_path):
        # 第一周导入
        base = _make_input(tmp_path, [_ts_row(ID="1", 工时ID="1", 工作日="2026-04-17")])
        Y.ingest(base)
        # 第二周:换一份只含当周数据的 xlsx,再导一次
        _write_timesheet(base, [_ts_row(ID="2", 工时ID="2", 工作日="2026-04-24")])
        r = Y.ingest(base)
        assert r["added"] == 1
        data = Y.build_yitian_data(base)
        assert data["meta"]["rows"] == 2                    # 两周累积
        assert data["meta"]["periodStart"] == "2026-04-17"
        assert data["meta"]["periodEnd"] == "2026-04-24"

    def test_reimport_same_week_updates_not_duplicates(self, tmp_path):
        base = _make_input(tmp_path, [_ts_row(ID="1", 工时ID="1", 工作成果="旧文本" + GOOD)])
        Y.ingest(base)
        _write_timesheet(base, [_ts_row(ID="1", 工时ID="1", 工作成果="新文本" + GOOD)])
        r = Y.ingest(base)
        assert r == {"added": 0, "updated": 1, "total": 1}
        data = Y.build_yitian_data(base)
        assert data["meta"]["rows"] == 1                    # 不变双份

    def test_build_on_empty_store_returns_none(self, tmp_path):
        (tmp_path / "input" / "yitian").mkdir(parents=True)
        (tmp_path / "data").mkdir()
        assert Y.build_yitian_data(str(tmp_path)) is None

    def test_missing_wid_column_rejected(self, tmp_path):
        # 缺 工时ID 列 → 缺列校验拦下,返回 None(不阻断主管线)
        base = _make_input(tmp_path, [_ts_row()], drop_cols=["工时ID"])
        assert Y.ingest(base) is None
