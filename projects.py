# projects.py
"""项目主域(Phase P)构建:三输入文件摄取 → 筛三部 → 售前映射 → 回款/成本聚合 → 健康度 + 质量。
镜像 pmis.py 模式:纯函数为主,文件读取集中,任一输入缺失优雅降级(不抛错、不阻断回款主流程)。"""
from __future__ import annotations

import os
import re
from typing import Any, Dict, List, Optional, Tuple

import config
from pmis import parse_pmis_money, parse_pmis_pct, read_pmis_sheet


def _open_workbook(path: str):
    """打开 xlsx;文件缺失/损坏返回 None。不用 read_only(WPS 导出 dimension 不可靠会截断行)。"""
    if not os.path.exists(path):
        return None
    import openpyxl
    try:
        return openpyxl.load_workbook(path, data_only=True)
    except Exception:
        return None


def _read_header_sheet(path: str, key_header: str) -> List[Dict[str, Any]]:
    """在所有 sheet 中找首行含 key_header 的表(跳过透视杂表),转 list[dict]。找不到返回 []。"""
    wb = _open_workbook(path)
    if wb is None:
        return []
    try:
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [str(h).strip() if h is not None else "" for h in rows[0]]
            if key_header not in headers:
                continue
            out = []
            for raw in rows[1:]:
                d = {}
                for i, h in enumerate(headers):
                    if h:
                        d[h] = raw[i] if i < len(raw) else None
                if any(v is not None for v in d.values()):
                    out.append(d)
            return out
        return []
    finally:
        wb.close()


def read_org_names(path: str) -> Tuple[set, set, int]:
    """组织架构表 → (姓名集合, L4组织集合, 行数)。按"表头含工号"自动选 sheet。
    若存在 新L3组织 列,仅收 交付实施三部 行(防误放全公司清单);姓名/L4 先 strip 再判空。"""
    rows = _read_header_sheet(path, "工号")
    if rows and any(r.get("新L3组织") for r in rows):
        rows = [r for r in rows if str(r.get("新L3组织") or "").strip() == config.DEPT_L3]
    names: set = set()
    l4s: set = set()
    for r in rows:
        name = str(r.get("姓名") or "").strip()
        if name:
            names.add(name)
        l4 = str(r.get("新L4组织") or "").strip()
        if l4:
            l4s.add(l4)
    return names, l4s, len(rows)



def read_mapping(path: str) -> List[Dict[str, str]]:
    """A.xlsx(无表头):A列=当前项目号 B列=桥接负责人 C列=已关闭项目号。AC 全有才收。"""
    wb = _open_workbook(path)
    if wb is None:
        return []
    try:
        ws = wb.worksheets[0]
        out = []
        for raw in ws.iter_rows(values_only=True):
            cur = str(raw[0]).strip() if raw and raw[0] is not None else ""
            owner = str(raw[1]).strip() if raw and len(raw) > 1 and raw[1] is not None else ""
            closed = str(raw[2]).strip() if raw and len(raw) > 2 and raw[2] is not None else ""
            if cur and closed:
                out.append({"current": cur, "owner": owner, "closed": closed})
        return out
    finally:
        wb.close()


def read_delivery(path: str) -> List[Dict[str, Any]]:
    """delivery_analysis 表:csv 优先(R1 起),缺失回退旧 xlsx;xlsx 按"表头含项目编号"自动选 sheet。"""
    if path.endswith(".csv"):
        from profit import read_csv_rows
        rows = read_csv_rows(path)
        if rows:
            return rows
        legacy = os.path.join(os.path.dirname(path), config.DELIVERY_FILE_LEGACY)
        return _read_header_sheet(legacy, "项目编号")
    return _read_header_sheet(path, "项目编号")


def read_top1000(path: str) -> Dict[str, Dict[str, str]]:
    """TOP1000.xlsx → {客户名称: {"level": 客户级别, "quad": 象限}}。
    复用 _read_header_sheet(找含"客户名称"表头的 sheet);缺文件/无表头 → {}(降级)。
    客户名称为空的行跳过;级别/象限 strip。"""
    rows = _read_header_sheet(path, "客户名称")
    out: Dict[str, Dict[str, str]] = {}
    for r in rows:
        name = str(r.get("客户名称") or "").strip()
        if not name:
            continue
        out[name] = {
            "level": str(r.get("客户级别") or "").strip(),
            "quad": str(r.get("象限") or "").strip(),
        }
    return out


def delivery_costs_for(row: Dict[str, Any]) -> List[Dict[str, Any]]:
    """delivery_analysis 一行 → 7 类目成本四元组(缺列降 None)。"""
    out = []
    for cat in config.DELIVERY_COST_CATEGORIES:
        out.append({
            "类别": cat,
            "预算金额": parse_pmis_money(row.get(f"{cat}_预算金额")),
            "实际发生": parse_pmis_money(row.get(f"{cat}_实际发生")),
            "剩余预算": parse_pmis_money(row.get(f"{cat}_剩余预算")),
            "消耗率": parse_pmis_pct(row.get(f"{cat}_消耗率")),
        })
    return out


def delivery_overspend_cats(delivery_costs: List[Dict[str, Any]]) -> List[str]:
    """交付费用超支类目(S1):实际发生 > 预算金额 的类目名(预算缺失不判)。"""
    out = []
    for c in delivery_costs or []:
        b, a = c.get("预算金额"), c.get("实际发生")
        if b is not None and a is not None and a > b:
            out.append(str(c.get("类别") or ""))
    return out


def payment_ratio_from_records(records_total: Optional[float], contract: Optional[float],
                               closed_contract: Optional[float]) -> Optional[float]:
    """回款完成率新口径(S1):流水累计 ÷ 合同总额(本项目优先,售前回退原项目)。
    分母缺失/0 → None(前端显 '-');无流水但有合同 → 0。"""
    denom = contract if contract else closed_contract
    if not denom or denom <= 0:
        return None
    return round((records_total or 0) / denom, 4)


def build_payment_summary(contract, nodes, pay_record):
    """系统核心口径回款摘要:计划侧=收款阶段节点(由 collection_stages 构建,含 status/reached);
    实际侧=项目流水(不分摊节点)。fromOrigin 由调用方写。"""
    actual_total = (pay_record or {}).get("total")
    return {
        "contract": contract,
        "actualTotal": actual_total,
        "paymentCount": (pay_record or {}).get("count", 0),
        "expectedTotal": round(sum(n["expectedPayment"] for n in nodes), 2),
        "nodeCount": len(nodes),
        "reachedCount": sum(1 for n in nodes if n["reached"]),
        "delayedCount": sum(1 for n in nodes if n["status"] == "延期"),
        "lastPaymentDate": (pay_record or {}).get("lastDate", ""),
        "fromOrigin": False,
    }


def aggregate_payment_pmis(nodes: List[Dict[str, Any]]) -> Dict[str, Any]:
    """项目回款子域聚合(收款阶段节点级,3E-3);形态同旧 payment 以兼容前端消费方。"""
    exp = sum(float(n.get("expectedPayment") or 0) for n in nodes)
    act = sum(float(n.get("receivedAmount") or 0) for n in nodes)
    rem = sum(float(n.get("unpaidAmount") or 0) for n in nodes)
    delayed = sum(1 for n in nodes if n.get("status") == "延期")
    return {
        "relatedNodeCount": len(nodes),
        "expectedTotal": round(exp, 2),
        "actualTotal": round(act, 2),
        "remainingTotal": round(rem, 2),
        "paymentRatio": None,
        "delayedCount": delayed,
    }


def compute_health(pm: Dict[str, Any], delayed_count: int) -> Dict[str, Any]:
    """四维三态健康度(spec 4.6;阈值集中在此,后续可调)。"""
    _ms = str(pm.get("progress", {}).get("里程碑进度状态") or "")
    progress_ab = any(k in _ms for k in config.MILESTONE_DELAYED_KEYWORDS)
    risk = pm.get("risk", {})
    risk_ab = (risk.get("最高等级") == "高") and ((risk.get("未关闭风险数") or 0) > 0)
    cost = pm.get("cost", {})
    ratio = cost.get("消耗比")
    cost_ab = bool(cost.get("项目超支")) or (ratio is not None and ratio > 1)
    pay_ab = delayed_count > 0
    n = sum([progress_ab, risk_ab, cost_ab, pay_ab])
    overall = "健康" if n == 0 else ("关注" if n == 1 else "风险")
    return {"progressAbnormal": progress_ab, "riskAbnormal": risk_ab,
            "costAbnormal": cost_ab, "paymentAbnormal": pay_ab, "overall": overall}


def parse_presale_customer_from_name(name) -> str:
    """从售前项目名解析客户:`售前服务-客户名称-12位数字` → 客户名称。
    贪婪 + 尾部数字锚定(客户名内含 '-' 也正确);不匹配 → ''。"""
    m = re.match(r'^' + re.escape(config.PRESALE_PREFIX) + r'-(.+)-(\d+)$', str(name or '').strip())
    return m.group(1).strip() if m else ''


def effective_customer(is_presale: bool, own_fc: str, orig_fc: str, project_name) -> str:
    """有效客户(单一来源):非售前=本项目最终客户;售前=原项目最终客户,空则项目名解析。"""
    if not is_presale:
        return own_fc
    if orig_fc:
        return orig_fc
    return parse_presale_customer_from_name(project_name)


def build_projects(project_pmis: Dict[str, Dict[str, Any]], org_names: set, org_l4s: set,
                   mapping: List[Dict[str, str]], delivery_rows: List[Dict[str, Any]],
                   top1000_map: Optional[Dict[str, Dict[str, str]]] = None) -> List[Dict[str, Any]]:
    """项目主表:PMIS 在建 → 筛三部(空人员清单=不过滤,降级) → 挂映射/成本/健康度/TOP1000。
    payment 字段由 preprocess 9f 循环用 aggregate_payment_pmis 填入。
    matched=False 守卫为防御性分支(现行 _assemble 恒 matched=True),供未来非 PMIS 来源项目使用。"""
    top1000_map = top1000_map or {}
    delivery_by_pid: Dict[str, Dict[str, Any]] = {}
    for r in delivery_rows:
        pid = str(r.get("项目编号") or "").strip()
        if pid:
            delivery_by_pid.setdefault(pid, r)
    map_by_current = {m["current"]: m for m in mapping}

    out = []
    for pid, pm in project_pmis.items():
        if pm.get("source") != "在建":
            continue
        team = pm.get("team", {})
        manager = str(team.get("项目经理") or "").strip()
        if org_names and manager not in org_names:
            continue
        drow = delivery_by_pid.get(pid)
        name = str(team.get("项目名称") or "").strip()
        if not name and drow:
            name = str(drow.get("项目名称") or "").strip()
        m = map_by_current.get(pid)
        # paymentAbnormal 暂用 0 计算，后续 9f 用收款阶段 delayed 重算
        health = (compute_health(pm, 0) if pm.get("matched")
                  else {"progressAbnormal": False, "riskAbnormal": False, "costAbnormal": False,
                        "paymentAbnormal": False, "overall": "无数据"})
        customer = pm.get("customer") or {}
        is_presale = ((pm.get("status") or {}).get("项目类型") == config.PRESALE_PROJECT_TYPE)
        related_closed = (m["closed"] if m else "")
        # 有效客户(单一来源):非售前=本项目最终客户;售前=原项目最终客户,空则项目名解析。
        # 用于 TOP1000 判定 + 落 Project.customer(前端各客户列/筛选统一读)。
        own_fc = str(customer.get("最终客户") or "").strip()
        orig_fc = str(((project_pmis.get(related_closed) or {}).get("customer") or {}).get("最终客户") or "").strip()
        final_customer = effective_customer(is_presale, own_fc, orig_fc, name)
        t1 = top1000_map.get(final_customer) if final_customer else None
        top1000 = "是" if (t1 and t1.get("level") == config.TOP1000_LEVEL) else "否"
        quadrant = (t1.get("quad") if t1 else "") or ""
        out.append({
            "projectId": pid,
            "projectName": name,
            "projectManager": manager,
            "orgL4": str(team.get("L4部门") or "").strip(),
            "orgL3_1": str(team.get("L3_1部门") or "").strip(),
            "合同编号": str(customer.get("合同编号") or "").strip(),
            "isPresale": is_presale,
            "relatedClosedId": related_closed,
            "deliveryCosts": delivery_costs_for(drow) if drow else [],
            "health": health,
            "top1000": top1000,
            "quadrant": quadrant,
            "customer": final_customer,
        })
    out.sort(key=lambda p: p["projectId"])
    return out


def count_closed_dept(pmis_dir: str, org_names: set) -> int:
    """已关闭 ∩ 交付三部 计数:项目中心-已关闭.xlsx 中 项目经理 ∈ org_names 的项目数。无人员清单→0。"""
    if not org_names:
        return 0
    rows = read_pmis_sheet(os.path.join(pmis_dir, config.PMIS_FILES_CLOSED["center"]))
    return sum(1 for r in rows if str(r.get("项目经理") or "").strip() in org_names)


def compute_projects_quality(projects: List[Dict[str, Any]],
                             project_pmis: Dict[str, Dict[str, Any]],
                             org_names: set, org_l4s: set, org_rows: int,
                             mapping: List[Dict[str, str]],
                             delivery_rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    """项目主域质量(spec 3.6):三文件记分卡 + 人员↔项目双向告警 + 售前映射覆盖。"""
    proj_ids = {p["projectId"] for p in projects}
    managers = {p["projectManager"] for p in projects if p["projectManager"]}
    staff_no_project = sorted(org_names - managers)
    manager_not_in_org = []
    if org_names:
        for pid, pm in project_pmis.items():
            if pm.get("source") != "在建" or pid in proj_ids:
                continue
            team = pm.get("team", {})
            mgr = str(team.get("项目经理") or "").strip()
            l4 = str(team.get("L4部门") or "").strip()
            if l4 in org_l4s and mgr:
                manager_not_in_org.append({
                    "projectId": pid,
                    "projectName": str(team.get("项目名称") or ""),
                    "manager": mgr,
                })
    presale = [p for p in projects if p["isPresale"]]
    presale_unmapped = [p for p in presale if not p["relatedClosedId"]]
    mapping_matched = sum(1 for m in mapping if m["current"] in proj_ids)
    delivery_matched = sum(1 for r in delivery_rows
                           if str(r.get("项目编号") or "").strip() in proj_ids)

    def stat(provided: bool, rows: int, matched: int) -> Dict[str, Any]:
        return {"provided": provided, "rows": rows, "matched": matched,
                "matchRate": round(matched / rows, 4) if rows else 0.0}

    return {
        "deptProjectCount": len(projects),
        "orgFile": stat(bool(org_names), org_rows, len(org_names & managers)),
        "mappingFile": stat(bool(mapping), len(mapping), mapping_matched),
        "deliveryFile": stat(bool(delivery_rows), len(delivery_rows), delivery_matched),
        "staffNoProject": [{"name": n} for n in staff_no_project],
        "managerNotInOrg": sorted(manager_not_in_org, key=lambda x: x["projectId"]),
        "presaleTotal": len(presale),
        "presaleMapped": len(presale) - len(presale_unmapped),
        "presaleUnmapped": [{"projectId": p["projectId"], "projectName": p["projectName"]}
                            for p in presale_unmapped],
    }


def load_dept_projects(input_dir: str, project_pmis: Dict[str, Dict[str, Any]],
                       mapping: List[Dict[str, str]] = None,
                       ) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """读组织架构+delivery → build_projects + 质量。mapping 由调用方先读(9a 也要用)。"""
    if mapping is None:
        mapping = []
    names, l4s, org_rows = read_org_names(os.path.join(input_dir, config.ORG_FILE))
    delivery = read_delivery(os.path.join(input_dir, config.DELIVERY_FILE))
    top1000 = read_top1000(os.path.join(input_dir, config.TOP1000_FILE))
    projects = build_projects(project_pmis, names, l4s, mapping, delivery, top1000)
    quality = compute_projects_quality(projects, project_pmis, names, l4s, org_rows,
                                       mapping, delivery)
    quality["closedDeptCount"] = count_closed_dept(
        os.path.join(input_dir, config.PMIS_DIRNAME), names)
    return projects, quality
